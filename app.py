import secrets
import hashlib
from datetime import datetime, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
    PSYCOPG2_AVAILABLE = True
except ImportError:
    psycopg2 = None
    RealDictCursor = None
    PSYCOPG2_AVAILABLE = False
from flask import Flask, request, jsonify, send_file, render_template, redirect
import pandas as pd, re, io, tempfile, os, json, copy, random, string, time, zipfile
from datetime import datetime
from email.mime.text import MIMEText
from openpyxl import load_workbook, Workbook
from zoneinfo import ZoneInfo

DATABASE_URL = os.environ.get('DATABASE_URL', '')

def get_db():
    if not PSYCOPG2_AVAILABLE:
        raise RuntimeError("psycopg2 not available")
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor, sslmode='require')
    return conn

def init_db():
    if not PSYCOPG2_AVAILABLE or not DATABASE_URL:
        print('psycopg2 or DATABASE_URL not available, skipping DB init')
        return
    try:
        conn = get_db()
        cur  = conn.cursor()
        # ... rest stays the same
        cur.execute('''
            CREATE TABLE IF NOT EXISTS activity_logs (
                id      SERIAL PRIMARY KEY,
                ts      TIMESTAMP DEFAULT NOW(),
                email   TEXT,
                action  TEXT,
                details TEXT
            )
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS sessions (
                token     TEXT PRIMARY KEY,
                email     TEXT,
                expiry    TIMESTAMP,
                last_seen TIMESTAMP
            )
        ''')
        cur.execute('''
            ALTER TABLE sessions ADD COLUMN IF NOT EXISTS last_seen TIMESTAMP
        ''')
        # One-time cleanup: any session whose last_seen was backfilled by the
        # column migration above (or never set) has no real "active" signal —
        # treat it as stale so it doesn't falsely show up as an online user.
        cur.execute('''
            UPDATE sessions SET last_seen = NULL
            WHERE last_seen IS NOT NULL AND last_seen < NOW() - INTERVAL '1 minute'
        ''')
        cur.execute('''
            CREATE TABLE IF NOT EXISTS otps (
                email  TEXT PRIMARY KEY,
                otp    TEXT,
                expiry TIMESTAMP
            )
        ''')
        conn.commit()
        cur.close()
        conn.close()
        print('DB initialized OK')
    except Exception as e:
        print(f'DB init error: {e}')

app = Flask(__name__, template_folder='templates')

# ── Deferred globals ─────────────────────────────────────────
SUBTYPE_HEADER_ROW = {}
SUBTYPE_MAP = {}
PV_LIST = []
CE_SUBTYPE_HEADER_ROW = {}
CE_SUBTYPE_MAP = {}
CE_PV_LIST = []
AP_SUBTYPE_HEADER_ROW = {}
AP_SUBTYPE_HEADER_ROW = {}
AP_SUBTYPE_MAP = {}
AP_PV_LIST = []
AP_PV_SUBCATEGORY = {}   # PV name -> title Sub Category (e.g. "TopWears", "InnerWears")
AP_SUBTYPE_SOURCE_PATH = {}  # subtype -> (filepath, sheet_name)
DROPDOWN_MAP = {}
_initialized = False

# ── Lazy initialization ──────────────────────────────────────
def _init_app():
    global _initialized, SUBTYPE_HEADER_ROW, SUBTYPE_MAP, PV_LIST
    global CE_SUBTYPE_HEADER_ROW, CE_SUBTYPE_MAP, CE_PV_LIST
    global AP_SUBTYPE_HEADER_ROW, AP_SUBTYPE_MAP, AP_PV_LIST, AP_PV_SUBCATEGORY, DROPDOWN_MAP
    if _initialized:
        return
    _initialized = True

    if DATABASE_URL and PSYCOPG2_AVAILABLE:
        try:
            init_db()
        except Exception as e:
            print(f'DB init error (non-fatal): {e}')

    try:
        SUBTYPE_HEADER_ROW, SUBTYPE_MAP = _build_header_row_map()
    except Exception as e:
        print(f"Warning: Could not build header row map: {e}")
        SUBTYPE_HEADER_ROW, SUBTYPE_MAP = {}, {}

    try:
        PV_LIST = load_pv_list()
    except Exception as e:
        print(f"Warning: Could not load PV_LIST: {e}")
        PV_LIST = []

    try:
        CE_SUBTYPE_HEADER_ROW, CE_SUBTYPE_MAP = _build_ce_header_row_map()
    except Exception as e:
        print(f"Warning: Could not build CE header row map: {e}")
        CE_SUBTYPE_HEADER_ROW, CE_SUBTYPE_MAP = {}, {}

    try:
        CE_PV_LIST = load_ce_pv_list()
    except Exception as e:
        print(f"Warning: Could not load CE_PV_LIST: {e}")
        CE_PV_LIST = []

    try:
        AP_SUBTYPE_HEADER_ROW, AP_SUBTYPE_MAP = _build_ap_header_row_map()
    except Exception as e:
        print(f"Warning: Could not build AP header row map: {e}")
        AP_SUBTYPE_HEADER_ROW, AP_SUBTYPE_MAP = {}, {}

    try:
        AP_PV_LIST, AP_PV_SUBCATEGORY = load_ap_pv_list()
    except Exception as e:
        print(f"Warning: Could not load AP PV list: {e}")
        AP_PV_LIST, AP_PV_SUBCATEGORY = [], {}

    try:
        DROPDOWN_MAP = _load_dropdown_map()
    except Exception as e:
        print(f"Warning: Could not load DROPDOWN_MAP: {e}")
        DROPDOWN_MAP = {}

@app.before_request
def before_request():
    _init_app()

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# ── Logging ────────────────────────────────────────────────────
def write_log(email, action, details=''):
    entry = {
        'ts':      datetime.utcnow().isoformat() + 'Z',
        'email':   email or 'anonymous',
        'action':  action,
        'details': details,
    }
    print(json.dumps(entry))
    if not DATABASE_URL: return
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute(
            'INSERT INTO activity_logs (email, action, details) VALUES (%s, %s, %s)',
            (email or 'anonymous', action, details)
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f'write_log error: {e}')

def read_logs(limit=200):
    if not DATABASE_URL: return []
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute(
            'SELECT ts, email, action, details FROM activity_logs ORDER BY ts DESC LIMIT %s',
            (limit,)
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return [{'ts': str(r['ts']), 'email': r['email'],
                 'action': r['action'], 'details': r['details']} for r in rows]
    except Exception as e:
        print(f'read_logs error: {e}')
        return []

# ── Load embedded template file once at startup ────────────────
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'Logic___Template_File.xlsx')

def _build_header_row_map():
    wb  = load_workbook(TEMPLATE_PATH)
    ws  = wb['PV Template']
    hdr_map    = {}
    static_map = {}
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == 'Category *':
            for r2 in range(r + 1, min(r + 5, ws.max_row + 1)):
                d2 = ws.cell(r2, 4).value
                if d2 and str(d2).strip() not in ('SubType','Category *','nan',''):
                    st = str(d2).strip()
                    if st not in hdr_map:
                        hdr_map[st] = r
                        hdrs = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                        entry = {}
                        for ci, col in enumerate(hdrs):
                            if col in ('Category *','SubCategory *','CategoryType *',
                                       'SubType','PVID *','discoveryCategoryIds','ProductCode *'):
                                v = str(ws.cell(r2, ci + 1).value or '').strip()
                                if v and v not in ('nan','NaN','None'):
                                    entry[col] = v
                        static_map[st] = entry
                    break
    return hdr_map, static_map

    print(f"Warning: Could not build header row map: {e}")
    SUBTYPE_HEADER_ROW, SUBTYPE_MAP = {}, {}

def load_pv_list():
    try:
        pv  = pd.read_excel(TEMPLATE_PATH, sheet_name='PV List')
        col = pv.columns[0]
        return [str(v).strip() for v in pv[col].dropna() if str(v).strip() not in ('nan','SubType')]
    except Exception as e:
        print(f"Warning: Could not load PV List: {e}")
        return []

    print(f"Warning: Could not load PV_LIST: {e}")
    PV_LIST = []

def get_template_wb_for_subtype(subtype):
    try:
        wb_src  = load_workbook(TEMPLATE_PATH)
        ws_src  = wb_src['PV Template']
        hdr_row = SUBTYPE_HEADER_ROW.get(subtype, 1)
        headers = [ws_src.cell(hdr_row, c).value for c in range(1, ws_src.max_column + 1)]
        while headers and headers[-1] is None:
            headers.pop()
    except Exception as e:
        print(f"Warning: Could not load template for {subtype}: {e}")
    headers = apply_header_renames(headers)
    wb_new       = Workbook()
    ws_new       = wb_new.active
    ws_new.title = 'PV Template'
    for ci, h in enumerate(headers, 1):
        ws_new.cell(1, ci).value = h
    apply_dropdown_validations(ws_new, headers, DROPDOWN_MAP)   # ← new line    
    return wb_new, headers

from openpyxl.worksheet.datavalidation import DataValidation

DROPDOWN_REF_PATH = os.path.join(os.path.dirname(__file__), 'Dropdown_Reference.xlsx')

def _load_dropdown_map(path=DROPDOWN_REF_PATH):
    try:
        if not os.path.exists(path):
            print(f"Warning: dropdown reference file not found at {path}")
            return {}
        wb = load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        dd_map = {}
        for r in range(1, ws.max_row + 1):
            key  = ws.cell(r, 1).value
            vals = ws.cell(r, 2).value
            if not key or not vals:
                continue
            key_str = str(key).strip()
            if key_str.upper() in ('ATTRIBUTE NAME', 'ATTRIBUTEKEY', 'ATTRIBUTE_NAME'):
                continue
            key_norm = key_str.upper()
            values = [v.strip() for v in str(vals).split(',') if v.strip()]
            if values:
                dd_map[key_norm] = values
        return dd_map
    except Exception as e:
        print(f"Warning: could not load dropdown reference from {path}: {e}")
        return {}

def apply_dropdown_validations(ws, headers, dd_map, max_row=1000):
    if not dd_map:
        return
    wb = ws.parent
    helper_name = '_DropdownLists'
    helper_ws = wb[helper_name] if helper_name in wb.sheetnames else wb.create_sheet(helper_name)
    helper_ws.sheet_state = 'hidden'
    next_col = helper_ws.max_column + 1 if helper_ws.max_row > 1 else 1

    for ci, h in enumerate(headers, 1):
        if not h:
            continue
        key_norm = str(h).replace('*', '').strip().upper()
        values = dd_map.get(key_norm)
        if not values:
            continue

        inline = ','.join(values)
        if len(inline) <= 255:
            dv = DataValidation(type='list', formula1=f'"{inline}"', allow_blank=True)
        else:
            col_letter = helper_ws.cell(1, next_col).column_letter
            for i, v in enumerate(values, start=1):
                helper_ws.cell(i, next_col).value = v
            ref = f"'{helper_name}'!${col_letter}$1:${col_letter}${len(values)}"
            next_col += 1
            dv = DataValidation(type='list', formula1=ref, allow_blank=True)

        target_col = ws.cell(1, ci).column_letter
        dv.add(f'{target_col}2:{target_col}{max_row}')
        ws.add_data_validation(dv)
# ── Default config ─────────────────────────────────────────────
DEFAULT_CONFIG = {
    "brands":              {},
    "biz_cat_id":          "BCAT-139461",
    "biz_cat_name":        "Footwear",
    "relationship":        "Parent",
    "catalog_status":      "ACTIVE",
    "status_remark":       "Ready to Launch",
    "tax_master_status":   "active",
    "gst_cgst":            50,
    "gst_sgst":            50,
    "gst_igst":            0,
    "country_of_origin":   "India",
    "product_condition":   "Fresh",
    "manufacturing_year":  "2026",
    "discovery_cat":       "DISCAT-135542",
}

CONFIG_PATH    = '/tmp/fillforge_config.json'
CE_CONFIG_PATH = '/tmp/fillforge_ce_config.json'
AP_CONFIG_PATH = '/tmp/fillforge_ap_config.json'

def _load_config(path, defaults):
    cfg = {k: v for k, v in defaults.items()}
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                saved = json.load(f)
            cfg.update(saved)
        except Exception as e:
            print(f'Warning: could not load config from {path}: {e}')
    return cfg

def _save_config(path, cfg):
    try:
        tmp = path + '.tmp'
        with open(tmp, 'w', encoding='utf-8') as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
        os.replace(tmp, path)
    except Exception as e:
        print(f'Warning: could not save config to {path}: {e}')

def get_config():
    return _load_config(CONFIG_PATH, DEFAULT_CONFIG)

def get_ce_config_from_disk():
    return _load_config(CE_CONFIG_PATH, CE_DEFAULT_CONFIG)

def get_ap_config_from_disk():
    return _load_config(AP_CONFIG_PATH, AP_DEFAULT_CONFIG)

config = get_config()

SMTP_HOST     = 'smtp.gmail.com'
SMTP_PORT     = 587
SMTP_USER     = os.environ.get('SMTP_USER', 'fillforgeotp@gmail.com')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', 'qfpyihawrqtchqwz')
OTP_EXPIRY_MINUTES = 10
SESSION_EXPIRY_DAYS = 7

# ── Utility ────────────────────────────────────────────────────
def safe(val, default=''):
    try:
        if pd.isna(val): return default
    except: pass
    s = str(val).strip() if val is not None else default
    return default if s in ('nan','None','NaN') else s

def parse_gst_percentage(val, default=5):
    """Handles GST values that may come in as 18, '18', '18%', ' 18 %', '18.0%' etc."""
    if val is None:
        return default
    s = str(val).strip()
    if s in ('', 'nan', 'None', 'NaN'):
        return default
    s = s.replace('%', '').replace(',', '').strip()
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return default

def detect_col(df, candidates):
    for c in candidates:
        if c in df.columns: return c
    for c in df.columns:
        for cand in candidates:
            if cand.lower() in c.lower(): return c
    return None

def build_col_map(df, hints):
    return {k: detect_col(df, v) for k, v in hints.items() if detect_col(df, v)}

def title_case_color(raw):
    if not raw: return raw
    return ' '.join(w.capitalize() for w in str(raw).strip().split())

def merge_colors(primary, secondary):
    p = title_case_color(safe(primary))
    s = title_case_color(safe(secondary))
    if not s or s in ('nan','None',''): return p
    if p.lower() == s.lower(): return p
    return f"{p} & {s}"

def extract_article(sku):
    s = str(sku).strip()
    if '_' in s:
        return s.split('_')[0].strip()
    COLOR_CODES = [
        'LBLUE','LGREY','LGRAY','LGREEN','LPINK','LBROWN',
        'DBLUE','DGREY','DGRAY','DGREEN','DPINK','DBROWN',
        'ONION','IVORY','ASSORTED','MULTI',
        'MRN','OLV','BGE','BRN','CRM','GLD','SLV','PPL','NVY',
        'MUL','AST','CLR','BLK','WHT','BLU','GRN','GRY','PNK',
        'ORG','PRP','BK','WH','BL','RD','GR','NV','PK','YL','OR',
    ]
    code_pattern = '(' + '|'.join(COLOR_CODES) + ')$'
    COLOR_WORDS = (r'BLACK|WHITE|BLUE|RED|GREEN|GREY|GRAY|BEIGE|BROWN|NAVY|PINK|YELLOW|'
                   r'ORANGE|PURPLE|PEACH|CREAM|MAROON|GOLD|SILVER|ASSORTED|MULTI|ONION|IVORY')
    cleaned = re.sub(r'[\s_]+[\dXx]+[-]+[\dXx]+[-]+[\dXx]*[\s\-]*(' + COLOR_WORDS + r')\b.*', '', s, flags=re.IGNORECASE).strip()
    cleaned = re.sub(r'[\s\-]+(' + COLOR_WORDS + r')\b.*', '', cleaned, flags=re.IGNORECASE).strip()
    cleaned = re.sub(r'\s*\(\d+.*\)$', '', cleaned).strip()
    cleaned = re.sub(r'[\s\-]+$', '', cleaned).strip()
    cleaned = re.sub(r'\s*-*\s*', '-', cleaned)
    if cleaned and cleaned != s:
        return cleaned
    m = re.search(code_pattern, s, re.IGNORECASE)
    if m:
        stripped = s[:m.start()].rstrip('-_').strip()
        if stripped:
            return stripped
    return s

def expand_size_range(size_str, size_type='UK'):
    s = str(size_str).strip()
    if ',' in s: return [x.strip() for x in s.split(',') if x.strip()]
    if '-' in s:
        parts = [x.strip() for x in s.split('-') if x.strip() and re.match(r'^\d+$', x.strip())]
        if len(parts) > 1:
            return parts
    m = re.match(r'^(\d+)[Xx](\d+)$', s)
    if m:
        start, end = int(m.group(1)), int(m.group(2))
        pfx = f"{size_type} " if size_type else ''
        if start > end:
            return [f"{pfx}{i}" for i in list(range(start, 14)) + list(range(1, end + 1))]
        return [f"{pfx}{i}" for i in range(start, end + 1)]
    if ' ' in s: return [x.strip() for x in s.split() if x.strip()]
    return [s] if s else []

def build_set_details(sizes_list, set_details_raw):
    raw = str(set_details_raw).strip() if set_details_raw else ''
    full = re.findall(r'((?:UK\s*)?\d+)\s*/\s*(\d+)', raw)
    if full:
        det   = [f'{s}/{q}' for s, q in full]
        # Changed '{q} pcs' to '{q}pcs'
        desc  = [f'{q}pcs of {s}' for s, q in full]
        avail = ', '.join(s for s, _ in full)
        return ', '.join(det), ', '.join(desc), avail
        
    dash = re.findall(r'((?:UK\s*)?\d+)\s*[-–]+\s*(\d+)', raw)
    if dash:
        det   = [f'{s}/{q}' for s, q in dash]
        # Changed '{q} pcs' to '{q}pcs'
        desc  = [f'{q}pcs of {s}' for s, q in dash]
        avail = ', '.join(s for s, _ in dash)
        return ', '.join(det), ', '.join(desc), avail
        
    qty_parts = [x.strip() for x in raw.split(',') if x.strip()]
    if qty_parts and sizes_list and all(q.isdigit() for q in qty_parts):
        if len(qty_parts) == len(sizes_list):
            pairs = list(zip(sizes_list, qty_parts))
        else:
            qty   = sum(int(q) for q in qty_parts) // max(len(sizes_list), 1)
            pairs = [(s, str(qty)) for s in sizes_list]
        det  = [f'{s}/{q}' for s, q in pairs]
        # Changed '{q} pcs' to '{q}pcs'
        desc = [f'{q}pcs of {s}' for s, q in pairs]
        return ', '.join(det), ', '.join(desc), ', '.join(s for s, _ in pairs)
        
    if sizes_list:
        det  = [f'{s}/1' for s in sizes_list]
        # Changed '1 pcs' to '1pcs'
        desc = [f'1pcs of {s}' for s in sizes_list]
        return ', '.join(det), ', '.join(desc), ', '.join(sizes_list)
        
    return raw, raw, ''

def parse_lbh(dim_str):
    parts = re.split(r'[Xx×]', str(dim_str).strip())
    if len(parts) == 3:
        try: return int(parts[0]), int(parts[1]), int(parts[2])
        except: pass
    return None, None, None

def derive_gender(subtype):
    st = str(subtype).lower()
    if "women" in st: return "Women's"
    if "men" in st:   return "Men's"
    if "girl" in st:  return "Girl's"
    if "boy" in st:   return "Boy's"
    if "infant" in st: return "Infant's"
    return ""

def normalize_df_columns(df):
    df.columns = [re.sub(r'\s+', ' ', str(c)).strip() for c in df.columns]
    return df

def make_title(brand, gender, upper, closure, fw_type, color):
    parts = [p for p in [brand, gender, upper, closure, fw_type] if p]
    base  = ' '.join(parts)
    return f"{base}, {color}" if color else base

def make_internal_title(brand, article, gender, upper, closure, fw_type, color, set_count, set_details_tpl):
    parts = [p for p in [brand, article, gender, upper, closure, fw_type] if p]
    base  = ' '.join(parts)
    return f"{base}, {color}, Set of {set_count} ({set_details_tpl})"

def make_description(brand, article, gender, upper, closure, fw_type, sole, color, sizes, set_count):
    title_part = ' '.join(p for p in [brand, gender, upper, closure, fw_type] if p)
    upper_l    = upper.lower()   if upper   else 'upper'
    sole_l     = sole.lower()    if sole    else 'sole'
    closure_l  = closure.lower() if closure else 'slip-on'
    return (
        f"Step out in style with the {title_part} ({article}). "
        f"The {upper_l} upper offers a snug, comfortable fit, while the {sole_l} sole delivers "
        f"reliable grip and cushioned support. "
        f"The {closure_l} closure makes wearing easy. "
        f"Color: {color}. Available sizes: {sizes}. "
        f"Set of {set_count} bulk-pack — ideal for retailers and resellers."
    )

def normalize_brands(brands_data):
    if not brands_data:
        return {}
    if isinstance(brands_data, dict):
        return brands_data
    if isinstance(brands_data, list):
        result = {}
        for item in brands_data:
            if isinstance(item, dict):
                name = item.get('name', item.get('brandName', ''))
                bid  = item.get('id',   item.get('brandId',   ''))
                if name:
                    result[name] = bid
            elif isinstance(item, (list, tuple)) and len(item) >= 2:
                result[item[0]] = item[1]
        return result
    return {}

def get_brand_info(drow, col_map, brands_dict):
    brands_dict = normalize_brands(brands_dict)
    if not brands_dict:
        return '', ''
    fallback_brand, fallback_id = next(iter(brands_dict.items()))
    brand_col = col_map.get('brand')
    if not brand_col:
        return fallback_brand, fallback_id
    try:
        file_brand = str(drow.get(brand_col, '')).strip()
    except:
        file_brand = ''
    if not file_brand or file_brand.lower() in ('nan', 'none', '', 'null'):
        return fallback_brand, fallback_id
    file_brand_lower = file_brand.lower().strip()
    for b_name, b_id in brands_dict.items():
        if b_name.lower().strip() == file_brand_lower:
            return b_name, b_id
    for b_name, b_id in brands_dict.items():
        bn = b_name.lower().strip()
        if bn in file_brand_lower or file_brand_lower in bn:
            return b_name, b_id
    file_first_word = file_brand_lower.split()[0] if file_brand_lower.split() else ''
    if file_first_word:
        for b_name, b_id in brands_dict.items():
            cfg_first_word = b_name.lower().strip().split()[0] if b_name.strip().split() else ''
            if cfg_first_word and cfg_first_word == file_first_word:
                return b_name, b_id
    return fallback_brand, fallback_id

# Header text to rename in the OUTPUT file only.
# Template files keep their original header text; only the generated
# output workbook shows the new name. Add more entries here any time
# an output column name needs to change, without touching template files.
HEADER_RENAME_MAP = {
    'PACKAGING_TYPE': 'PACK_TYPE *',
}

def _norm_header(h):
    """Strip a trailing ' *' so 'PACKAGING_TYPE' and 'PACKAGING_TYPE *' match as the same header."""
    if h is None:
        return h
    return re.sub(r'\s*\*\s*$', '', str(h)).strip()

def apply_header_renames(headers):
    renamed = []
    for h in headers:
        if not h:
            renamed.append(h)
            continue
        norm = _norm_header(h)
        renamed.append(HEADER_RENAME_MAP.get(norm, h))
    return renamed
    
# ── Column hints ────────────────────────────────────────────────
DUMP_COL_HINTS = {
    'sku':            ['Seller SKU ID','Seller SKU_ID','ChildSKU *','ChildSKU','SKU'],
    'article':        ['Article Code','Article Number','ARTICLE_NUMBER'],
    'image':          ['Image Links','Image Link','ImageURL1','imageURL1 *'],
    'vertical':       ['Product Vertical','Subtype','SubType','PVName'],
    'gender':         ['Gender','GENDER *','GENDER'],
    'fw_type':        ['Foot Wear Type','FW Type','Footwear Type','FOOTWEAR_TYPE *'],
    'upper_material': ['Upper Material','UPPER_MATERIAL *'],
    'closure_type':   ['Closure Type','CLOSURE_TYPE *'],
    'sole_material':  ['Sole Material','SOLE_MATERIAL *'],
    'set_of':         ['Set of','Set Of','*Quantity','Set Count','SET_COUNT *','set count','Quantity'],
    'size_type':      ['Size Type','Unit of Measurement','Unit of Measuremen'],
    'color':          ['Primary Colour','Product Primary Color','Product Color','PRODUCT_COLOR *'],
    'color2':         ['Secondary Colour','Secondary Color'],
    'sizes':          ['Available Sizes','AVAILABLE_SIZES *'],
    'set_details':    ['Set Details','SET_DETAILS *'],
    'heel_height':    ['Heel Height','Heel_Height','HEEL_HEIGHT'],
    'heel_type':      ['Heel Type','Heel_Type','HEEL_TYPE'],
    'hsn':            ['HSN','*HSN Code','HSN Code','hsnCode *'],
    'gst':            ['GSTpercentage','*GST','GST','gstPercentage *'],
    'moq':            ['*MOQ','MOQ *','MOQ'],
    'mrp':            ['*MRP full Set','MRP *','MRP full Set','MRP'],
    'sp':             ['*Selling Price Full Set','SellingPrice *','*Selling Price per Pair'],
    'weight':         ['*Product Weight (In KG) Full Ste','PRODUCT_WEIGHT_IN_KG *'],
    'dims':           ['*Product Dimension (LXBXH) Full Set','Product Dimension'],
    'dim_uom':        ['*Product Dimension UOM','PRODUCT_DIMENSION_UOM *'],
    'packing':        ['Packing Type','PACKAGING_TYPE *'],
    'country':        ['Country of Origin','COUNTRY_OF_ORIGIN *'],
    'product_desc':   ['Product Description','productDescription *'],
    'brand':          ['Brand','Brand Name','brandName *','brand_name'],
}

BASE_COL_HINTS = {
    'article': ['Article Number','Article Code','ARTICLE_NUMBER'],
    'sku':     ['Seller SKU ID','Seller SKU_ID','ChildSKU'],
}

def fill_template(ws, headers, rows_df, col_map, subtype, existing_articles, existing_skus):
    tcol = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        tcol[h] = i + 1
        tcol.setdefault(_norm_header(h), i + 1)
    _cfg        = get_config()
    brands_dict = normalize_brands(_cfg.get('brands', {}))
    fallback_brand, fallback_id = ('', '')
    if brands_dict:
        fallback_brand, fallback_id = next(iter(brands_dict.items()))
    gender  = derive_gender(subtype)
    st_data = SUBTYPE_MAP.get(subtype, {})
    skipped, filled = [], []

    for _, drow in rows_df.iterrows():
        brand, brand_id = get_brand_info(drow, col_map, brands_dict)
        if not brand and fallback_brand:
            brand    = fallback_brand
            brand_id = fallback_id

        sku_raw = safe(drow.get(col_map.get('sku',''), ''))
        art_raw = safe(drow.get(col_map.get('article',''), ''))
        article = art_raw if (art_raw and '_' not in art_raw) else extract_article(sku_raw)

        if article.upper() in existing_articles or sku_raw.upper() in existing_skus:
            skipped.append({'sku': sku_raw, 'article': article, 'reason': 'Already exists in base data'})
            continue

        filled_count  = len(filled) + 1
        row_idx  = filled_count + 1

        size_type   = safe(drow.get(col_map.get('size_type',''), 'UK')) or 'UK'
        sizes_raw   = safe(drow.get(col_map.get('sizes',''), ''))
        set_det_raw = safe(drow.get(col_map.get('set_details',''), ''))
        _sc_raw     = safe(drow.get(col_map.get('set_of',''), '0'))
        _sc_nums    = re.findall(r'\d+', str(_sc_raw))
        set_count   = int(_sc_nums[0]) if _sc_nums else 0
        color       = merge_colors(
            drow.get(col_map.get('color',''), ''),
            drow.get(col_map.get('color2',''), '')
        )
        mrp         = drow.get(col_map.get('mrp',''), '')
        sp          = drow.get(col_map.get('sp',''), '')
        weight      = drow.get(col_map.get('weight',''), '')
        dim_raw     = safe(drow.get(col_map.get('dims',''), ''))
        hsn         = drow.get(col_map.get('hsn',''), '')
        gst         = drow.get(col_map.get('gst',''), 5)
        moq         = drow.get(col_map.get('moq',''), 1)
        upper_mat   = safe(drow.get(col_map.get('upper_material',''), ''))
        sole_mat    = safe(drow.get(col_map.get('sole_material',''), ''))
        closure     = safe(drow.get(col_map.get('closure_type',''), ''))
        heel_ht     = safe(drow.get(col_map.get('heel_height',''), ''))
        heel_type   = safe(drow.get(col_map.get('heel_type',''), ''))
        img_url     = safe(drow.get(col_map.get('image',''), ''))
        packing     = safe(drow.get(col_map.get('packing',''), '')) or 'Loose Packing'
        country     = safe(drow.get(col_map.get('country',''), '')) or _cfg['country_of_origin']
        dim_uom     = safe(drow.get(col_map.get('dim_uom',''), '')) or 'cm'
        fw_type     = safe(drow.get(col_map.get('fw_type',''), ''))
        prod_desc   = safe(drow.get(col_map.get('product_desc',''), ''))

        for v, field in [(packing,'packing'),(country,'country')]:
            if v in ('nan','None',''):
                if field == 'country': country = _cfg['country_of_origin']
                if field == 'packing': packing = 'Loose Packing'

        sizes_list = expand_size_range(sizes_raw, size_type)
        set_details_tpl, set_desc, avail_sizes = build_set_details(sizes_list, set_det_raw)
        if not avail_sizes: avail_sizes = ', '.join(sizes_list)

        title          = make_title(brand, gender, upper_mat, closure, fw_type, color)
        internal_title = make_internal_title(brand, article, gender, upper_mat, closure, fw_type, color, set_count, set_details_tpl)
        description    = prod_desc if prod_desc else make_description(brand, article, gender, upper_mat, closure, fw_type, sole_mat, color, avail_sizes, set_count)

        try:    mrp    = float(mrp)      if str(mrp).strip()    not in ('','nan') else ''
        except: mrp    = ''
        try:    sp     = float(sp)       if str(sp).strip()     not in ('','nan') else ''
        except: sp     = ''
        try:    hsn    = int(float(hsn)) if str(hsn).strip()    not in ('','nan') else ''
        except: hsn    = ''
        try:    gst    = parse_gst_percentage(gst, default=5)
        except: gst    = 5
        try:    moq    = int(float(moq))
        except: moq    = 1
        try:    weight = float(weight)   if str(weight).strip() not in ('','nan') else ''
        except: weight = ''
        L, B, H = parse_lbh(dim_raw)

        product_code = article

        row_data = {
            'Category *':                                  st_data.get('Category *', 'Footwear'),
            'SubCategory *':                               st_data.get('SubCategory *', ''),
            'CategoryType *':                              st_data.get('CategoryType *', ''),
            'SubType':                                     subtype,
            'PVID *':                                      st_data.get('PVID *', ''),
            'BusinessCategoryId *':                        _cfg['biz_cat_id'],
            'BusinessCategoryName *':                      _cfg['biz_cat_name'],
            'ProductCode *':                               product_code,
            'Relationship *':                              _cfg['relationship'],
            'ParentProductId *':                           sku_raw,
            'ChildSKU *':                                  sku_raw,
            'MRP *':                                       mrp,
            'SellingPrice *':                              sp,
            'MOQ *':                                       moq,
            'title *':                                     title,
            'internalTitle *':                             internal_title,
            'brandId *':                                   brand_id,
            'brandName *':                                 brand,
            'imageURL1 *':                                 img_url,
            'catalogStatus *':                             _cfg['catalog_status'],
            'statusRemark':                                _cfg['status_remark'],
            'discoveryCategoryIds':                        st_data.get('discoveryCategoryIds', _cfg['discovery_cat']),
            'productDescription *':                        description,
            'PRODUCT_IDENTIFIER *':                        'Set',
            'SET_NAME *':                                  f'Set of {set_count}',
            'SET_COUNT *':                                 set_count,
            'PACK_NAME *':                                 'Pack of 1',
            'PACK_OF *':                                   1,
            'IS_COMBO *':                                  'yes',
            'AVAILABLE_SIZES *':                           avail_sizes,
            'SET_DETAILS *':                               set_details_tpl,
            'SET_DESCRIPTION *':                           set_desc,
            'PRODUCT_COLOR *':                             color,
            'ARTICLE_NUMBER *':                            article,
            'MODEL_NAME *':                                article,
            'PRODUCT_CONDITION *':                         _cfg['product_condition'],
            'UNIT_OF_MEASUREMENT_SINGULAR *':              'Pair',
            'UNIT_OF_MEASUREMENT_PLURAL *':                'Pairs',
            'UNIT_OF_MEASUREMENT_SINGULAR_ABBREVIATION *': 'Pair',
            'UNIT_OF_MEASUREMENT_PLURAL_ABBREVIATION *':   'Pairs',
            'SELLER_SKU_ID *':                             sku_raw,
            'PACKAGING_TYPE *':                            packing,
            'GENDER *':                                    gender,
            'AGE_GROUP *':                                 '',
            'CLOSURE_TYPE *':                              closure,
            'COUNTRY_OF_ORIGIN *':                         country,
            'MANUFACTURING_YEAR':                          _cfg['manufacturing_year'],
            'PRODUCT_LENGTH *':                            L,
            'PRODUCT_BREADTH *':                           B,
            'PRODUCT_HEIGHT *':                            H,
            'PRODUCT_DIMENSION_UOM *':                     dim_uom,
            'PRODUCT_WEIGHT_IN_KG *':                      weight,
            'FOOTWEAR_TYPE *':                             fw_type,
            'HEEL_HEIGHT':                                 heel_ht,
            'HEEL_HEIGHT *':                               heel_ht,
            'HEEL_TYPE':                                   heel_type,
            'HEEL_TYPE *':                                 heel_type,
            'SOLE_MATERIAL *':                             sole_mat,
            'UPPER_MATERIAL *':                            upper_mat,
            'hsnCode *':                                   hsn,
            'gstPercentage *':                             gst,
            'cgstShare *':                                 _cfg['gst_cgst'],
            'sgstShare *':                                 _cfg['gst_sgst'],
            'igstShare *':                                 _cfg['gst_igst'],
            'taxMasterStatus':                             _cfg['tax_master_status'],
        }

        for col_name, val in row_data.items():
            if col_name in tcol and val is not None and str(val) not in ('None',''):
                ws.cell(row=row_idx, column=tcol[col_name]).value = val

        filled.append({'sku': sku_raw, 'article': article})

    return len(filled), skipped


# ═══════════════════════════════════════════════════════════════
# CONSUMER ELECTRONICS MODULE
# ═══════════════════════════════════════════════════════════════

CE_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'Consumer_Electronics_Template.xlsx')

def _build_ce_header_row_map():
    wb  = load_workbook(CE_TEMPLATE_PATH)
    ws  = wb['CE - PV Template']
    hdr_map    = {}
    static_map = {}
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == 'Category *':
            for r2 in range(r + 1, min(r + 5, ws.max_row + 1)):
                # Per spec: map by Column D = SubType (fall back to CategoryType only if SubType is blank)
                c4_val      = ws.cell(r2, 4).value
                c3_val      = ws.cell(r2, 3).value
                subtype_val = c4_val if c4_val else c3_val
                if subtype_val and str(subtype_val).strip() not in ('CategoryType *','SubType','Category *','nan',''):
                    st = str(subtype_val).strip()
                    if st not in hdr_map:
                        hdr_map[st] = r
                        hdrs  = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                        entry = {}
                        for ci, col in enumerate(hdrs):
                            if col in ('Category *','SubCategory *','CategoryType *',
                                       'SubType','PVID *','discoveryCategoryIds'):
                                v = str(ws.cell(r2, ci + 1).value or '').strip()
                                if v and v not in ('nan','NaN','None'):
                                    entry[col] = v
                        static_map[st] = entry
                    break
    return hdr_map, static_map

    print(f"Warning: Could not build CE header row map: {e}")
    CE_SUBTYPE_HEADER_ROW, CE_SUBTYPE_MAP = {}, {}

def load_ce_pv_list():
    try:
        pv  = pd.read_excel(CE_TEMPLATE_PATH, sheet_name='Category List')
        col = pv.columns[0]
        return [str(v).strip() for v in pv[col].dropna()
                if str(v).strip() not in ('nan','CategoryType *','')]
    except Exception as e:
        print(f"Warning: Could not load CE Category List: {e}")
        return []

    print(f"Warning: Could not load CE_PV_LIST: {e}")
    CE_PV_LIST = []

CE_DEFAULT_CONFIG = {
    "brands":              {},
    "biz_cat_id":          "BCAT-139438",
    "biz_cat_name":        "Consumer Electronics",
    "relationship":        "Parent",
    "catalog_status":      "ACTIVE",
    "status_remark":       "Ready to Launch",
    "tax_master_status":   "active",
    "gst_cgst":            50,
    "gst_sgst":            50,
    "gst_igst":            0,
    "country_of_origin":   "India",
    "product_condition":   "Fresh",
    "manufacturing_year":  "2026",
    "discovery_cat":       "DISCAT-135528",
}

CE_DUMP_COL_HINTS = {
    'sku':            ['Child SKU','ChildSKU *','ChildSKU','SKU','SKU ID','Seller SKU ID'],
    'article':        ['Model Number','MODEL NUMBER','Model NUMBER','Model Name',
                       'Name of the model/Title name',
                       'Article Number','Article Code','ARTICLE_NUMBER'],
    'image':          ['Main Image URL','Image Links','Image Link','ImageURL1','imageURL1 *'],
    'image2':         ['Other Image URL 1','Other Image URL1','imageURL2'],
    'image3':         ['Other Image URL 2','Other Image URL2','imageURL3'],
    'image4':         ['Other Image URL 3','Other Image URL3','imageURL4'],
    'image5':         ['Other Image URL 4','Other Image URL4','imageURL5'],
    'image6':         ['Other Image URL 5','Other Image URL5','imageURL6'],
    'vertical':       ['Product Type','Sub Type','Product Sub-type','CategoryType *','Subtype','SubType'],
    'brand':          ['Brand','Brand Name','brandName *','brand_name'],
    'mrp':            ['MRP','*MRP full Set','MRP *','MRP full Set'],
    'sp':             ['Selling Price','SellingPrice *','*Selling Price per Pair'],
    'moq':            ['*Minimum Order Quantity','*MOQ','MOQ *','MOQ'],
    'color':          ['Product Color','Product Colour','Colour','Primary Colour','PRODUCT_COLOR *'],
    'product_desc':   ['Product Description','productDescription *'],
    'hsn':            ['HSN Code','HSN','*HSN Code','hsnCode *'],
    'gst':            ['GST','*GST','gstPercentage *'],
    'weight':         ['Product Weight','Product Weight (KG)','*Product Weight (In KG) Full Ste','PRODUCT_WEIGHT_IN_KG *'],
    'dims':           ['*Product Dimension (LXBXH)','Product Dimension (LXBXH) Full Set','Product Dimension'],
    'dim_uom':        ['*Product Dimension UOM','PRODUCT_DIMENSION_UOM *'],
    'packing':        ['Packaging Type','PACKAGING_TYPE *'],
    'country':        ['Country/Region of Origin','Country of Origin','COUNTRY_OF_ORIGIN *'],
    'warranty':       ['Warranty Period','Warranty'],
    'battery':        ['Battery Capacity','BATTERY_CAPACITY_MAH *'],
    'charging_type':  ['Charging type supported','CHARGING_TYPE_SUPPORTED *'],
    'ram':            ['RAM','RAM *'],
    'storage':        ['Storage Capacity','Internal Storage','INTERNAL_STORAGE *'],
    'sim_type':       ['Sim Type','SIM_TYPE *'],
    'os':             ['Operating System','OPERATING_SYSTEM_OS *'],
    'os_version':     ['Operating System','OPERATING_SYSTEM_OS *'],
    'front_camera':   ['Front Camera','FRONT_CAMERA_RESOLUTION *'],
    'back_camera':    ['Back Camera','PRIMARY_CAMERA_RESOLUTION *'],
    'screen_size':    ['Screen Size','Display Size','DISPLAY_SIZE *'],
    'display_type':   ['Display Type','DISPLAY_TYPE *'],
    'processor_core': ['Processor Core','NUMBER_OF_PROCESSOR_CORES *'],
    'network_support':['Network Support','Network'],
    'bluetooth':      ['Bluetooth Version','BLUETOOTH_VERSION *'],
    'product_type':   ['Product Type','Product Sub-type'],
    # ── Extra attributes required by per-SubType title formulas ───────────
    'output_voltage':         ['Output Voltage','Output Current','OUTPUT_CURRENT_OR_VOLTAGE *','Output Voltage/Current'],
    'adapter_connector_type': ['Adapter Connector Type','ADAPTER_CONNECTOR_TYPE *','Connector Type','Adapter\nConnector\nType'],
    'battery_type':           ['Battery Type','BATTERY_TYPE','Battery type'],
    'speaker_type':           ['Speaker Type','SPEAKER_TYPE *'],
    'compatible_brand':       ['Compatible Brand','COMPATIBLE_BRAND *','Compatible With Brand'],
    'compatible_model':       ['Compatible Brand + Model Name','Compatible Model','Compatible Model Number','COMPATIBLE_BRAND_MODEL *','Compatible Brand & Model'],
    'material':               ['Material','MATERIAL *','Product Material'],
    'case_cover_type':        ['Case Cover Type','Case & Cover Type','CASE_COVER_TYPE *'],
    'number_of_connectors':   ['Number of Connectors','No of Connectors','NUMBER_OF_CONNECTORS',
                               'Number of\nConnectors','Number\nof\nConnect\nors'],
    'rotation':               ['Rotation','Rotation/Adjustability','ROTATION_OR_ADJUSTABILITY *'],
    'screen_guard_type':      ['Screen Guard Type','Screen Guard / Protector Type','SCREEN_GUARD_OR_PROTECTOR_TYPE *'],
    'coverage':               ['Coverage','COVERAGE *'],
    'mic_type':               ['Mic Type','Microphone Type','MIC_TYPE *','MIC_TYPE'],
    'wired_or_wireless':      ['Wired or Unwired','Wired or Wireless','WIRED_OR_WIRELESS *'],
    'connector_type_for_wired': ['Connector type','Connector Type','CONNECTOR_TYPE_FOR_WIRED *'],
    'output_connector_type':  ['Output Connector Type','OUTPUT_CONNECTOR_TYPE','Connector Type',
                               'Connect\nor Type'],
    'number_of_output_ports': ['Number of Output Ports','Number of Output Port','No of Output Ports',
                               'NO_OF_OUTPUT_PORTS','Number of Output','Number\nof\nOutput'],
    'port_type':              ['Port type','Port Type','PORT_TYPE *','PORT_TYPE'],
    'number_of_ports':        ['Number of Ports','No of Ports','NO_OF_ADAPTER_PORTS *'],
    'memory_card_type':       ['Memory Card Type','MEMORY_CARD_TYPE *'],
    'storage_capacity':       ['Storage Capacity','Internal Storage','STORAGE_CAPACITY *','Memory Capacity'],
    'speed_class':            ['Speed Class','SPEED_CLASS *','Class'],
    'holder_type':       ['Holder Type','HOLDER_TYPE *','Type'],
    'product_condition': ['Product Condition','PRODUCT_CONDITION *','Condition'],
    'ticket_id':      ['Ticket ID','TicketID','Ticket_ID','ticket_id'],
    'cable_included': ['Cable Included','cable_included'],
    'cable_length':   ['Cable Length','Cable Length In Meter','CABLE_LENGTH_IN_METER *'],
    'set_includes':   ['Set Includes','SET_INCLUDES','PRODUCT_TYPE *'],
    'cable_type':     ['Cable Type','CABLE_TYPE'],
    'material':               ['Material','MATERIAL *','Product Material'],
    'case_cover_type':        ['Case Cover Type','Case & Cover Type','CASE_COVER_TYPE *'],
    'pattern':                ['Pattern','DESIGN *','Design'],
    'case_cover_closure':     ['Case & Cover Closure','Case Cover Closure','CLOSURE_TYPE *'],
    'thickness':              ['Screen Card Thickness','Thickness','THICKNESS *'],
}

CE_BASE_COL_HINTS = {
    'article': ['Model Number','MODEL NUMBER','Model Name','Name of the model/Title name',
                'Article Number','Article Code','ARTICLE_NUMBER'],
    'sku':     ['Child SKU','ChildSKU','SKU ID'],
}

def extract_model_name(title_name):
    s = str(title_name).strip()
    if not s: return s
    cleaned = re.sub(r'\s*\([^)]*\)\s*', ' ', s)
    cleaned = re.sub(r'\s+(Fresh|Seal Open|Non Activated|Open Seal)\s*$', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s+\d+\s*GB\s*RAM.*$', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s+\d+\s*GB\s*ROM.*$', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s+\d+G\s*.*$', '', cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r'\s+', ' ', cleaned).strip()
    return cleaned if cleaned else s

def make_ce_title(brand, model_name, back_camera, category_type, ram_storage, color, condition):
    core_parts = []
    if brand:         core_parts.append(brand)
    if model_name:    core_parts.append(model_name)
    if back_camera:   core_parts.append(f'{back_camera} Camera')
    if category_type: core_parts.append(category_type)
    base = ' '.join(core_parts)
    suffix_parts = []
    if ram_storage: suffix_parts.append(ram_storage)
    color_condition = ' '.join(p for p in [color, f'({condition})' if condition else ''] if p)
    if color_condition: suffix_parts.append(color_condition)
    if suffix_parts:
        return f"{base}, {', '.join(suffix_parts)}"
    return base

def make_feature_phone_title(brand, model_name, screen_size, category_type, color, condition):
    core_parts = []
    if brand:       core_parts.append(brand)
    if model_name:  core_parts.append(model_name)
    if screen_size: core_parts.append(f'{screen_size}" Display')
    if category_type: core_parts.append(category_type)
    base = ' '.join(core_parts)
    suffix_parts = []
    color_condition = ' '.join(p for p in [color, f'({condition})' if condition else ''] if p)
    if color_condition: suffix_parts.append(color_condition)
    if suffix_parts:
        return f"{base}, {', '.join(suffix_parts)}"
    return base

def extract_compatible_brand(text):
    """'Samsung S23' -> 'Samsung' (first word of the compatible model string)."""
    s = str(text).strip()
    return s.split()[0] if s else ''


# ═══════════════════════════════════════════════════════════════════
# Per-SubType title + internalTitle formulas (from CE - Mapping Logic)
# Each builder receives a `f` dict with the resolved field values and
# returns a clean, comma-normalised string.
# ═══════════════════════════════════════════════════════════════════
def _ce_join(parts, sep=' '):
    """Join non-empty stringified parts, collapsing whitespace."""
    out = [str(p).strip() for p in parts if p is not None and str(p).strip() not in ('','nan','None','NaN')]
    return sep.join(out)

def _ce_cond(cond):
    """Render product condition as '(Fresh)' or '' if blank."""
    c = str(cond).strip() if cond else ''
    return f'({c})' if c and c not in ('nan','None','NaN') else ''

def _ce_screen_size_for_title(val):
    """'6.77 inches' -> '6.77' — strips units, keeps the numeric value only.
    Used for the title's X" Display formatting. DISPLAY_SIZE * output keeps
    the raw input value untouched."""
    s = str(val).strip()
    if not s:
        return ''
    m = re.search(r'(\d+(?:\.\d+)?)', s)
    return m.group(1) if m else s

def _ce_compose(core_tokens, tail_tokens):
    """Compose 'core, tail1, tail2' — drops empty tails and the leading comma."""
    core = _ce_join(core_tokens)
    tails = [t for t in (_ce_join([x]) for x in tail_tokens) if t]
    return f"{core}, {', '.join(tails)}" if tails else core

# 1. Feature Phones
def title_feature_phones(f, internal=False):
    color_cond  = _ce_join([f.get('color'), _ce_cond(f.get('condition'))], ' ')
    subtype_str = str(f.get('subtype', '')).strip()
    if subtype_str.endswith('Phones'):
        subtype_str = subtype_str[:-1]
    return _ce_compose(
        [f['brand'], f['model'], f'{f["screen_size_title"]}" Display' if f.get('screen_size_title') else '', subtype_str],
        [color_cond],
    )

# 2. Smart Phone
def title_smart_phone(f, internal=False):
    ram_rom = _ce_join([f.get('ram'), f.get('storage')], ' + ')
    color_cond = _ce_join([f.get('color'), _ce_cond(f.get('condition'))], ' ')
    return _ce_compose(
        [f['brand'], f['model'], f'{f["back_camera"]} Camera' if f.get('back_camera') else '', f['subtype']],
        [ram_rom, color_cond],
    )

# 3. Mobile Adapters & Cables
def title_mobile_adapters(f, internal=False):
    cable_suffix = 'With 1m Cable' if str(f.get('cable_included', '')).strip().lower() == 'yes' else ''
    core = [f['brand'], f.get('condition')]
    if internal:
        core.append(f.get('model'))
    core += [f.get('output_voltage'), f.get('adapter_connector_type'), 'Adapter']
    if cable_suffix:
        core.append(cable_suffix)
    return _ce_compose(core, [f.get('color')])

# 4. Hair Trimmer
def title_hair_trimmer(f, internal=False):
    return _ce_compose(
        [f['brand'], f['model'], f.get('condition'), f.get('battery_type'), f['subtype']],
        [f['color']],
    )

# 5. Speakers
def title_speakers(f, internal=False):
    base = _ce_join([f['brand'], f['model'], f.get('condition'), f.get('speaker_type')])
    if internal:
        return _ce_compose([base], [f['color']])
    return base

# 6. Mobile Case & Covers
def title_mobile_cases(f, internal=False):
    compat = f.get('compatible_model')  # full "Compatible Brand + Model Name" value
    core = [f['brand'], f.get('condition')]
    if internal:
        core.append(f.get('model'))     # Model Number
    core += [compat, f.get('material'), f.get('pattern'), f.get('case_cover_type')]
    return _ce_compose(core, [f['color']])

# 7. Earphones
def title_earphones(f, internal=False):
    base = _ce_join([f['brand'], f['model'], f.get('condition'), 'Wired Earphones'])
    if internal:
        return _ce_compose([base], [f['color']])
    return base

# 8. Headsets
def title_headsets(f, internal=False):
    base = _ce_join([f['brand'], f['model'], f.get('condition'), f['subtype']])
    if internal:
        return _ce_compose([base], [f['color']])
    return base

# 9. Memory Cards
def title_memory_cards(f, internal=False):
    return _ce_compose(
        [f['brand'], f.get('storage_capacity') or f.get('storage'), f.get('memory_card_type'), f['subtype']],
        [f.get('speed_class')],
    )

# 10. Mobile Cables
def title_mobile_cables(f, internal=False):
    core = [f['brand'], f.get('condition')]
    if internal:
        core.append(f.get('model'))
    core += [f.get('number_of_connectors'), f.get('adapter_connector_type'), f['subtype']]
    return _ce_compose(core, [f.get('color')])

# 11. Mobile Holders
def title_mobile_holders(f, internal=False):
    core = [f['brand'], f.get('condition')]
    if internal:
        core.append(f.get('model'))
    core += [f['subtype'], f.get('rotation')]
    return _ce_compose(core, [f['color']])

# 12. Screen Guards / Protectors
def title_screen_guards(f, internal=False):
    compat = f.get('compatible_model')  # full "Compatible Brand + Model Name" value, as-is
    core = [f['brand']]
    if internal:
        core.append(f.get('model'))     # Model Number
    core += [f.get('screen_guard_type'), f.get('coverage'), 'For', compat]
    return _ce_join(core)

# 13. Neck Bands
def title_neck_bands(f, internal=False):
    base = _ce_join([f['brand'], f['model'], f.get('condition'), 'Neckband'])
    if internal:
        return _ce_compose([base], [f['color']])
    return base

# 14. Microphone
def title_microphone(f, internal=False):
    return _ce_join([f['brand'], f['model'], f.get('condition'),
                     f.get('mic_type'), 'with', f.get('output_connector_type')])

# 15. Power Bank
def title_power_bank(f, internal=False):
    core = [f['brand'], f.get('condition')]
    if internal:
        core.append(f.get('model'))
    core += [f.get('battery'), f['subtype'], f.get('number_of_output_ports')]
    return _ce_compose(core, [f.get('color')])

# 16. Smart Watches
def title_smart_watches(f, internal=False):
    return _ce_compose(
        [f['brand'], f['model'], f.get('condition'),
         f'{f["screen_size_title"]}" Display' if f.get('screen_size_title') else '', f['subtype']],
        [f['color']],
    )

# 17. TWS Ear Buds
def title_tws_earbuds(f, internal=False):
    base = _ce_join([f['brand'], f['model'], f.get('condition'), 'Earbuds'])
    if internal:
        return _ce_compose([base], [f['color']])
    return base

# 18. Pendrive
def title_pendrive(f, internal=False):
    core = [f['brand']]
    if internal:
        core.append(f.get('model'))  # Model Number
    core += [f.get('condition'), f.get('storage_capacity'), f['subtype']]
    return _ce_compose(core, [f['color']])

# Dispatcher: subtype → builder function
CE_TITLE_BUILDERS = {
    'Feature Phones':              title_feature_phones,
    'Smart Phone':                 title_smart_phone,
    'Mobile Adapters & Cables':    title_mobile_adapters,
    'Hair Trimmer':                title_hair_trimmer,
    'Speakers':                    title_speakers,
    'Mobile Case & Covers':        title_mobile_cases,
    'Earphones':                   title_earphones,
    'Headsets':                    title_headsets,
    'Memory Cards':                title_memory_cards,
    'Mobile Cables':               title_mobile_cables,
    'Mobile Holders':              title_mobile_holders,
    'Screen Guards / Protectors':  title_screen_guards,
    'Neck Bands':                  title_neck_bands,
    'Microphone':                  title_microphone,
    'Power Bank':                  title_power_bank,
    'Smart Watches':               title_smart_watches,
    'TWS Ear Buds':                title_tws_earbuds,
    'Pendrive':                    title_pendrive,
}

def build_ce_titles(subtype, fields):
    """Return (title, internal_title) for a given SubType using its bespoke formula.
    Falls back to the generic builder for SubTypes without a formula
    (currently: Projectors, Soundbars, Webcams)."""
    builder = CE_TITLE_BUILDERS.get(subtype)
    if builder:
        try:
            return builder(fields, internal=False), builder(fields, internal=True)
        except Exception as e:
            print(f'Warning: per-SubType title builder failed for {subtype}: {e}')
    # Generic fallback — original behaviour
    ram_storage = _ce_join([fields.get('ram'), fields.get('storage')], '+')
    title = make_ce_title(fields.get('brand',''), fields.get('model',''),
                          fields.get('back_camera',''), fields.get('subtype',''),
                          ram_storage, fields.get('color',''), fields.get('condition',''))
    return title, title

def make_ce_description(brand, model_name, category_type, ram, storage, processor, battery,
                        screen_size, display_type, color, front_camera, back_camera, os):
    parts = []
    if brand and model_name:
        parts.append(f"Experience the {brand} {model_name}")
    if category_type:
        parts.append(f", a powerful {category_type}")
    if ram and storage:
        parts.append(f" featuring {ram} RAM and {storage} internal storage")
    if processor:
        parts.append(f", powered by a {processor} processor")
    if battery:
        parts.append(f". Equipped with a {battery} battery")
    if screen_size and display_type:
        parts.append(f", it boasts a {screen_size} {display_type} display")
    if front_camera and back_camera:
        parts.append(f". Capture stunning photos with {back_camera} rear and {front_camera} front cameras")
    if color:
        parts.append(f". Available in {color} color")
    if os:
        parts.append(f". Runs on {os}")
    return ''.join(parts) + ". Ideal for everyday use with reliable performance and modern features."

def extract_from_description(desc, field_type):
    if not desc: return ''
    desc_lower = str(desc).lower()
    if field_type == 'display_type':
        for dt in ['super amoled','amoled','pls lcd','lcd','ips','oled','tft']:
            if dt in desc_lower:
                if dt == 'super amoled': return 'Super AMOLED'
                if dt == 'pls lcd':      return 'PLS LCD'
                return dt.upper()
        return ''
    if field_type == 'charging_type':
        if 'usb type-c' in desc_lower or 'usb c' in desc_lower or 'type-c' in desc_lower:
            return 'USB Type-C'
        if 'micro usb' in desc_lower or 'micro-usb' in desc_lower:
            return 'Micro USB'
        if 'lightning' in desc_lower:
            return 'Lightning'
        return ''
    if field_type == 'bluetooth':
        m = re.search(r'bluetooth\s*(\d+\.?\d*)', desc_lower)
        return f'Bluetooth {m.group(1)}' if m else ''
    if field_type == 'processor':
        for proc in ['dimensity','snapdragon','helio','exynos','kirin','mediatek']:
            if proc in desc_lower:
                m = re.search(rf'{proc}\s+([a-z]*\d+\s*[a-z]*)', desc_lower, re.IGNORECASE)
                if m:
                    return f"{proc.capitalize()} {m.group(1).strip().upper()}"
                return proc.capitalize()
        return ''
    return ''

def get_ce_template_wb_for_subtype(subtype):
    try:
        wb_src  = load_workbook(CE_TEMPLATE_PATH)
        ws_src  = wb_src['CE - PV Template']
        hdr_row = CE_SUBTYPE_HEADER_ROW.get(subtype, 2)
        headers = [ws_src.cell(hdr_row, c).value for c in range(1, ws_src.max_column + 1)]
        while headers and headers[-1] is None:
            headers.pop()
    except Exception as e:
        print(f"Warning: Could not load CE template for {subtype}: {e}")
        headers = []
    headers = apply_header_renames(headers)
    wb_new       = Workbook()
    ws_new       = wb_new.active
    ws_new.title = 'CE - PV Template'
    for ci, h in enumerate(headers, 1):
        ws_new.cell(1, ci).value = h
    apply_dropdown_validations(ws_new, headers, DROPDOWN_MAP)   # ← new line    
    return wb_new, headers

def fill_ce_template(ws, headers, rows_df, col_map, subtype, existing_articles, existing_skus):
    tcol        = {h: i+1 for i, h in enumerate(headers) if h}
    _ce_cfg     = get_ce_config_from_disk()
    brands_dict = normalize_brands(_ce_cfg.get('brands', {}))
    fallback_brand, fallback_id = ('', '')
    if brands_dict:
        fallback_brand, fallback_id = next(iter(brands_dict.items()))
    st_data = CE_SUBTYPE_MAP.get(subtype, {})
    skipped, filled = [], 0

    for _, drow in rows_df.iterrows():
        brand, brand_id = get_brand_info(drow, col_map, brands_dict)
        if not brand and fallback_brand:
            brand    = fallback_brand
            brand_id = fallback_id

        sku_raw    = safe(drow.get(col_map.get('sku',''), ''))
        model_num  = safe(drow.get(col_map.get('article',''), ''))
        article    = model_num if model_num else sku_raw
        model_name = article

        if article.upper() in existing_articles or sku_raw.upper() in existing_skus:
            skipped.append({'sku': sku_raw, 'article': article, 'reason': 'Already exists in base data'})
            continue

        filled  += 1
        row_idx  = filled + 1

        mrp          = drow.get(col_map.get('mrp',''), '')
        sp           = drow.get(col_map.get('sp',''), '')
        moq          = drow.get(col_map.get('moq',''), 1)
        color        = safe(drow.get(col_map.get('color',''), ''))
        weight       = drow.get(col_map.get('weight',''), '')
        dim_raw      = safe(drow.get(col_map.get('dims',''), ''))
        hsn          = drow.get(col_map.get('hsn',''), '')
        gst          = drow.get(col_map.get('gst',''), 18)
        packing      = safe(drow.get(col_map.get('packing',''), '')) or 'BOX'
        country      = safe(drow.get(col_map.get('country',''), '')) or _ce_cfg['country_of_origin']
        dim_uom      = safe(drow.get(col_map.get('dim_uom',''), '')) or 'cm'
        prod_desc    = safe(drow.get(col_map.get('product_desc',''), ''))
        warranty     = safe(drow.get(col_map.get('warranty',''), ''))
        battery      = safe(drow.get(col_map.get('battery',''), ''))
        charging     = safe(drow.get(col_map.get('charging_type',''), ''))
        ram          = safe(drow.get(col_map.get('ram',''), ''))
        storage      = safe(drow.get(col_map.get('storage',''), ''))
        sim_type     = safe(drow.get(col_map.get('sim_type',''), ''))
        os_raw     = safe(drow.get(col_map.get('os',''), ''))
            # OPERATING_SYSTEM_OS * → just the text part e.g. "Android"
        os         = re.sub(r'\s*\d+.*$', '', os_raw).strip()
            # OS_VERSION * → full value e.g. "Android 15"
        os_version = os_raw
        front_cam    = safe(drow.get(col_map.get('front_camera',''), ''))
        back_cam     = safe(drow.get(col_map.get('back_camera',''), ''))
        screen_size  = safe(drow.get(col_map.get('screen_size',''), ''))
        display_type = safe(drow.get(col_map.get('display_type',''), ''))
        proc_core    = safe(drow.get(col_map.get('processor_core',''), ''))
        network      = safe(drow.get(col_map.get('network_support',''), ''))
        bluetooth    = safe(drow.get(col_map.get('bluetooth',''), ''))
        prod_type    = safe(drow.get(col_map.get('product_type',''), ''))
        wired_or_wireless        = safe(drow.get(col_map.get('wired_or_wireless',''), ''))
        connector_type_for_wired = safe(drow.get(col_map.get('connector_type_for_wired',''), ''))

        img_url  = safe(drow.get(col_map.get('image',''), ''))
        img2_url = safe(drow.get(col_map.get('image2',''), ''))
        img3_url = safe(drow.get(col_map.get('image3',''), ''))
        img4_url = safe(drow.get(col_map.get('image4',''), ''))
        img5_url = safe(drow.get(col_map.get('image5',''), ''))
        img6_url = safe(drow.get(col_map.get('image6',''), ''))

        desc = prod_desc if prod_desc else ''
        if not display_type:
            display_type = extract_from_description(desc, 'display_type')
        if not charging:
            charging = extract_from_description(desc, 'charging_type')
        if not bluetooth:
            bluetooth = extract_from_description(desc, 'bluetooth')

        ram_rom     = f"{ram} + {storage}" if (ram and storage) else ''
        _raw_cond = safe(drow.get(col_map.get('product_condition',''), ''))
        condition = _raw_cond if _raw_cond else ''

        # ── Per-SubType title + internalTitle (from CE - Mapping Logic) ──
        ce_fields = {
            'brand':                  brand,
            'model':                  model_name,
            'subtype':                subtype,
            'condition':              condition,
            'color':                  color,
            'ram':                    ram,
            'storage':                storage,
            'back_camera':            back_cam,
            'screen_size':            screen_size,
            'screen_size_title':      _ce_screen_size_for_title(screen_size),
            'battery':                battery,
            'output_voltage':         safe(drow.get(col_map.get('output_voltage',''), '')),
            'adapter_connector_type': safe(drow.get(col_map.get('adapter_connector_type',''), '')),
            'battery_type':           safe(drow.get(col_map.get('battery_type',''), '')),
            'speaker_type':           safe(drow.get(col_map.get('speaker_type',''), '')),
            'compatible_model':       safe(drow.get(col_map.get('compatible_model',''), '')),
            'compatible_brand':       (safe(drow.get(col_map.get('compatible_brand',''), ''))
                           or extract_compatible_brand(safe(drow.get(col_map.get('compatible_model',''), '')))),
            'pattern':                safe(drow.get(col_map.get('pattern',''), '')),
            'case_cover_closure':     safe(drow.get(col_map.get('case_cover_closure',''), '')),
            'material':               safe(drow.get(col_map.get('material',''), '')),
            'case_cover_type':        safe(drow.get(col_map.get('case_cover_type',''), '')),
            'number_of_connectors':   safe(drow.get(col_map.get('number_of_connectors',''), '')),
            'rotation':               safe(drow.get(col_map.get('rotation',''), '')),
            'screen_guard_type':      safe(drow.get(col_map.get('screen_guard_type',''), '')),
            'coverage':               safe(drow.get(col_map.get('coverage',''), '')),
            'mic_type':               safe(drow.get(col_map.get('mic_type',''), '')),
            'output_connector_type':  safe(drow.get(col_map.get('output_connector_type',''), '')),
            'number_of_output_ports': safe(drow.get(col_map.get('number_of_output_ports',''), '')),
            'port_type':              safe(drow.get(col_map.get('port_type',''), '')),
            'memory_card_type':       safe(drow.get(col_map.get('memory_card_type',''), '')),
            'storage_capacity':       safe(drow.get(col_map.get('storage_capacity',''), '')),
            'speed_class':            safe(drow.get(col_map.get('speed_class',''), '')),
            'holder_type':            safe(drow.get(col_map.get('holder_type',''), '')),
            'cable_included':         safe(drow.get(col_map.get('cable_included',''), '')),
            'thickness':              safe(drow.get(col_map.get('thickness',''), '')),
            
        }
        title, internal_title = build_ce_titles(subtype, ce_fields)

        description = prod_desc if prod_desc else make_ce_description(
            brand, model_name, subtype, ram, storage, proc_core, battery,
            screen_size, display_type, color, front_cam, back_cam, os
        )

        package_contents = 'Handset' if (prod_type and 'smart phone' in prod_type.lower()) else ''

        L, B, H = parse_lbh(dim_raw)

        weight_clean = ''
        if weight:
            m = re.search(r'(\d+\.?\d*)', str(weight))
            if m:
                try:
                    weight_clean = float(m.group(1))
                except (ValueError, TypeError):
                    weight_clean = ''
        try:    mrp = float(mrp)      if str(mrp).strip()    not in ('','nan') else ''
        except: mrp = ''
        try:    sp  = float(sp)       if str(sp).strip()     not in ('','nan') else ''
        except: sp  = ''
        try:    hsn = int(float(hsn)) if str(hsn).strip()    not in ('','nan') else ''
        except: hsn = ''
        gst = parse_gst_percentage(gst, default=18)
        try:    moq = int(float(moq)) if str(moq).strip() not in ('', 'nan') else 1
        except: moq = 1    

        row_data = {
            'Category *':                                  st_data.get('Category *', 'Consumer Electronics'),
            'SubCategory *':                               st_data.get('SubCategory *', 'Mobile'),
            'CategoryType *':                              subtype,
            'SubType':                                     subtype,
            'PVID *':                                      st_data.get('PVID *', ''),
            'BusinessCategoryId *':                        _ce_cfg['biz_cat_id'],
            'BusinessCategoryName *':                      _ce_cfg['biz_cat_name'],
            'Relationship *':                              _ce_cfg['relationship'],
            'ParentProductId *':                           sku_raw,
            'ChildSKU *':                                  sku_raw,
            'MRP *':                                       mrp,
            'SellingPrice *':                              sp,
            'MOQ *':                                       moq,
            'title *':                                     title,
            'internalTitle *':                             internal_title,
            'brandId *':                                   brand_id,
            'brandName *':                                 brand,
            'imageURL1 *':                                 img_url,
            'imageURL2':                                   img2_url,
            'imageURL3':                                   img3_url,
            'imageURL4':                                   img4_url,
            'imageURL5':                                   img5_url,
            'imageURL6':                                   img6_url,
            'catalogStatus *':                             _ce_cfg['catalog_status'],
            'statusRemark':                                _ce_cfg['status_remark'],
            'discoveryCategoryIds':                        st_data.get('discoveryCategoryIds', _ce_cfg['discovery_cat']),
            'productDescription *':                        description,
            'PRODUCT_IDENTIFIER *':                        'Set',
            'SET_NAME *':                                  'Set of 1',
            'SET_COUNT *':                                 1,
            'PACK_NAME *':                                 'Pack of 1',
            'PACK_OF *':                                   1,
            'IS_COMBO *':                                  'yes',
            'AVAILABLE_SIZES *':                           '1',
            'SET_DETAILS *':                                ram_rom,
            'SET_DESCRIPTION *':                           f'1pc of {subtype}',
            'PRODUCT_COLOR *':                             color,
            'ARTICLE_NUMBER *':                            article,
            'MODEL_NAME *':                                model_name,
            'PRODUCT_CONDITION *':                         condition,
            'UNIT_OF_MEASUREMENT_SINGULAR *':              'Piece',
            'UNIT_OF_MEASUREMENT_PLURAL *':                'Pieces',
            'UNIT_OF_MEASUREMENT_SINGULAR_ABBREVIATION *': 'Pc',
            'UNIT_OF_MEASUREMENT_PLURAL_ABBREVIATION *':   'Pcs',
            'SELLER_SKU_ID *':                             sku_raw,
            'PACKAGING_TYPE *':                            packing,
            'DESCRIPTION':                                 '',
            'BATTERY_CAPACITY_MAH *':                      battery,
            'CHARGING_TYPE_SUPPORTED *':                   charging,
            'COUNTRY_OF_ORIGIN *':                         country,
            # ── Adapter & Charger specific ──────────────────
            'ADAPTER_CONNECTOR_TYPE *':                    ce_fields.get('adapter_connector_type', ''),
            'OUTPUT_CURRENT_OR_VOLTAGE *':                 ce_fields.get('output_voltage', ''),
            'NO_OF_ADAPTER_PORTS *':                       safe(drow.get(col_map.get('number_of_ports',''), '')),
            'CABLE_LENGTH_IN_METER *':                     safe(drow.get(col_map.get('cable_length',''), '')),
            # ── Mobile Cable specific ───────────────────────
            'NUMBER_OF_CONNECTORS':                        ce_fields.get('number_of_connectors', ''),
            'CABLE_TYPE':                                  safe(drow.get(col_map.get('cable_type',''), '')),
            # -- Audio Device specific (Speakers, Earphones, Neck Bands, TWS Ear Buds) --
            # Speakers:     CONNECTOR_TYPE_FOR_WIRED *, SPEAKER_TYPE *, WARRANTY *, WIRED_OR_WIRELESS *
            # Earphones:    CONNECTOR_TYPE_FOR_WIRED *, MIC_TYPE *,     WARRANTY *, WIRED_OR_WIRELESS *
            # Neck Bands:   MIC_TYPE *, WARRANTY *, WIRED_OR_WIRELESS *
            # TWS Ear Buds: MIC_TYPE *, WARRANTY *, WIRED_OR_WIRELESS *
            'CONNECTOR_TYPE_FOR_WIRED *':                  connector_type_for_wired,
            'SPEAKER_TYPE *':                              ce_fields.get('speaker_type', ''),
            'WARRANTY *':                                  warranty,
            'WIRED_OR_WIRELESS *':                         wired_or_wireless,
            'MIC_TYPE *':                                  ce_fields.get('mic_type', ''),
            'MIC_TYPE':                                    ce_fields.get('mic_type', ''),
            # ── Power Bank & Adapter shared ─────────────────
            'PORT_TYPE *':                                 safe(drow.get(col_map.get('port_type',''), '')),
            'PRODUCT_TYPE *':                              safe(drow.get(col_map.get('set_includes',''), '')),
            # ── Power Bank specific ─────────────────────────
            'PACKAGING_CLASSIFICATION':                    packing,
            'EAN *':                                       '',
            'IMPORTED_BY':                                 '',
            'KEY_FEATURES':                                '',
            'MANUFACTURING_YEAR':                          _ce_cfg['manufacturing_year'],
            'PACKAGE_CONTENTS *':                          package_contents,
            'PORT_TYPE':                                   '',
            'PRODUCT_BREADTH *':                           B,
            'PRODUCT_DIMENSION_UOM *':                     dim_uom,
            'PRODUCT_HEIGHT *':                            H,
            'PRODUCT_LENGTH *':                            L,
            'PRODUCT_WEIGHT_IN_KG *':                      weight_clean,
            'PRODUCT_MANUFACTURING_CITY':                  '',
            'PRODUCT_MANUFACTURING_STATE':                 '',
            'WARRANTY':                                    warranty,
            'MANUFACTURER':                                '',
            'OPERATING_SYSTEM_OS *':                       os,
            'OS_VERSION *':                                os_version,
            'DISPLAY_SIZE *':                              screen_size,
            'DISPLAY_TYPE *':                              display_type,
            'DISPLAY_RESOLUTION':                          '',
            'RAM *':                                       ram,
            'INTERNAL_STORAGE *':                          storage,
            'EXPANDABLE_STORAGE *':                        '',
            'SIM_TYPE *':                                  sim_type,
            'BLUETOOTH_VERSION *':                         bluetooth,
            'EXPANDABLE_STORAGE_TYPE':                     '',
            'EXPANDABLE_STORAGE_CAPACITY_MAX':             '',
            'BATTERY_TYPE':                                '',
            'REMOVABLE_BATTERY':                           '',
            'HYBRID_SIM_SLOT':                             '',
            'NETWORK_TYPE_SUPPORTED':                      network,
            'AUDIO_JACK':                                  '',
            'FM_RADIO *':                                  '',
            'TORCH_OR_FLASHLIGHT *':                       '',
            'RAM_ROM *':                                   ram_rom,
            'PROCESSOR_BRAND_AND_MODEL_NAME':              extract_from_description(desc, 'processor'),
            'NUMBER_OF_PROCESSOR_CORES *':                 proc_core,
            'PRIMARY_CAMERA_RESOLUTION *':                 back_cam,
            'FRONT_CAMERA_RESOLUTION *':                   front_cam,
            'REAR_FLASH':                                  '',
            'SIM_SIZE':                                    '',
            'WIFI':                                        '',
            'FINGERPRINT_SENSOR':                          '',
            'CLOCK_SPEED':                                 '',
            'REFRESH_RATE':                                '',
            'TOUCHSCREEN_TYPE':                            '',
            'PRIMARY_CAMERA_SETUP':                        '',
            'FRONT_FLASH':                                 '',
            'VIDEO_RECORDING_RESOLUTION':                  '',
            'FAST_CHARGING_WATTAGE':                       '',
            'WIRELESS_CHARGING_SUPPORT':                   '',
            'GPS_SUPPORT':                                 '',
            'NFC_SUPPORT':                                 '',
            'INFRARED_IR_BLASTER':                         '',
            'FINGERPRINT_SENSOR_POSITION':                 '',
            'FACE_UNLOCK':                                 '',
            'WATER_RESISTANCE_RATING':                     '',
            'hsnCode *':                                   hsn,
            'gstPercentage *':                             gst,
            'cgstShare *':                                 _ce_cfg['gst_cgst'],
            'sgstShare *':                                 _ce_cfg['gst_sgst'],
            'igstShare *':                                 _ce_cfg['gst_igst'],
            'cess':                                        '',
            'sinTax':                                      '',
            'vatPercentage':                               '',
            'otherCess':                                   '',
            'validityPeriodStartDate':                     '',
            'validityPeriodEndDate':                       '',
            'declarationForm':                             '',
            'taxMasterStatus':                             _ce_cfg['tax_master_status'],
            'ProductCode *':                               model_num,
            'MATERIAL *':                                  ce_fields.get('material', ''),
            'COMPATIBLE_BRAND_MODEL *':                    ce_fields.get('compatible_model', ''),
            'CASE_COVER_TYPE *':                           ce_fields.get('case_cover_type', ''),
            'CLOSURE_TYPE *':                              ce_fields.get('case_cover_closure', ''),
            'DESIGN *':                                    ce_fields.get('pattern', ''),
            'COMPATIBLE_BRAND *':                          ce_fields.get('compatible_brand', ''),
            'COVERAGE *':                                  ce_fields.get('coverage', ''),
            'SCREEN_GUARD_OR_PROTECTOR_TYPE *':            ce_fields.get('screen_guard_type', ''),
            'THICKNESS *':                                 ce_fields.get('thickness', ''),
            'STORAGE_CAPACITY *':                          ce_fields.get('storage_capacity', ''),
            'CONNECTOR_TYPE *':                            ce_fields.get('output_connector_type', ''),
        }

        for col_name, val in row_data.items():
            if col_name in tcol and val is not None and str(val) not in ('None',''):
                ws.cell(row=row_idx, column=tcol[col_name]).value = val

    return filled, skipped

# ═══════════════════════════════════════════════════════════════
# APPAREL & FASHION MODULE
# ═══════════════════════════════════════════════════════════════

AP_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'Apparel_Fashion_Template.xlsx')
AP_BOTTOMWEAR_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'ApparelBottomWearTemplates.xlsx')
AP_TOPBOTTOMWEAR_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'Top&BottomWearTemplate.xlsx')

def _build_ap_header_row_map_from_path(path, sheet_name):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    hdr_map = {}
    static_map = {}
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == 'Category *':
            for r2 in range(r + 1, min(r + 5, ws.max_row + 1)):
                d2 = ws.cell(r2, 4).value
                if d2 and str(d2).strip() not in ('SubType', 'Category *', 'nan', ''):
                    st = str(d2).strip()
                    if st not in hdr_map:
                        hdr_map[st] = r
                        hdrs = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                        entry = {}
                        for ci, col in enumerate(hdrs):
                            if col in ('Category *', 'SubCategory *', 'CategoryType *',
                                       'SubType', 'PVID *', 'discoveryCategoryIds'):
                                v = str(ws.cell(r2, ci + 1).value or '').strip()
                                if v and v not in ('nan', 'NaN', 'None'):
                                    entry[col] = v
                        static_map[st] = entry
                    break
    return hdr_map, static_map


def _build_ap_header_row_map():
    global AP_SUBTYPE_SOURCE_PATH
    AP_SUBTYPE_SOURCE_PATH = {}

    hdr_map, static_map = _build_ap_header_row_map_from_path(
        AP_TEMPLATE_PATH, 'AF - PV Templates'
    )
    for st in hdr_map:
        AP_SUBTYPE_SOURCE_PATH[st] = (AP_TEMPLATE_PATH, 'AF - PV Templates')

    if os.path.exists(AP_BOTTOMWEAR_TEMPLATE_PATH):
        try:
            bw_hdr, bw_map = _build_ap_header_row_map_from_path(
                AP_BOTTOMWEAR_TEMPLATE_PATH, 'AF Bottomwear - Template'
            )
            for st in bw_hdr:
                hdr_map[st] = bw_hdr[st]
                static_map[st] = bw_map[st]
                AP_SUBTYPE_SOURCE_PATH[st] = (AP_BOTTOMWEAR_TEMPLATE_PATH, 'AF Bottomwear - Template')
        except Exception as e:
            print(f"Warning: Could not build bottomwear header row map: {e}")

    if os.path.exists(AP_TOPBOTTOMWEAR_TEMPLATE_PATH):
        try:
            tb_hdr, tb_map = _build_ap_header_row_map_from_path(
                AP_TOPBOTTOMWEAR_TEMPLATE_PATH, 'AF Top&BottomWear - Template'
            )
            for st in tb_hdr:
                hdr_map[st] = tb_hdr[st]
                static_map[st] = tb_map[st]
                AP_SUBTYPE_SOURCE_PATH[st] = (AP_TOPBOTTOMWEAR_TEMPLATE_PATH, 'AF Top&BottomWear - Template')
        except Exception as e:
            print(f"Warning: Could not build top&bottomwear header row map: {e}")

    return hdr_map, static_map
def load_ap_pv_list_from_path(path, pv_sheet_name):
    wb = load_workbook(path)
    ws = wb[pv_sheet_name]
    pv_list = []
    pv_subcat_map = {}
    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(1, col).value
        if not header_val:
            continue
        header_str = str(header_val).strip()
        if not header_str or header_str.lower() == 'nan':
            continue
        subcat_name = header_str.split('=', 1)[1].strip() if '=' in header_str else header_str
        if not subcat_name:
            continue
        for r in range(3, ws.max_row + 1):
            v = ws.cell(r, col).value
            if v is None:
                continue
            pv = str(v).strip()
            if not pv or pv.lower() in ('nan', 'none'):
                continue
            if pv not in pv_list:
                pv_list.append(pv)
            pv_subcat_map[pv] = subcat_name
    return pv_list, pv_subcat_map


def load_ap_pv_list():
    """
    Reads the 'Product Vertical List' sheet, which has ONE COLUMN PER SUB CATEGORY.
    Row 1  -> header cell like 'Sub Category=TopWears' / 'Sub Catgeory=InnerWears'
              (text after '=' is taken as the Sub Category name; typo-tolerant)
    Row 2  -> label row ('SubType=Product Vertical') - skipped
    Row 3+ -> Product Vertical names for that column's Sub Category

    Returns:
        pv_list       - flat, de-duplicated list of all PV names (for the dropdown)
        pv_subcat_map - {pv_name: sub_category_name}, used to route title formulas
    """
    pv_list, pv_subcat_map = load_ap_pv_list_from_path(
        AP_TEMPLATE_PATH, 'Product Vertical List'
    )
    if os.path.exists(AP_BOTTOMWEAR_TEMPLATE_PATH):
        try:
            bw_list, bw_map = load_ap_pv_list_from_path(
                AP_BOTTOMWEAR_TEMPLATE_PATH, 'Product Vertical List'
            )
            for pv in bw_list:
                if pv not in pv_list:
                    pv_list.append(pv)
                pv_subcat_map[pv] = bw_map.get(pv, 'BottomWear')
        except Exception as e:
            print(f"Warning: Could not load bottomwear PV list: {e}")

    if os.path.exists(AP_TOPBOTTOMWEAR_TEMPLATE_PATH):
        try:
            tb_list, tb_map = load_ap_pv_list_from_path(
                AP_TOPBOTTOMWEAR_TEMPLATE_PATH, 'Product Vertical List'
            )
            for pv in tb_list:
                if pv not in pv_list:
                    pv_list.append(pv)
                pv_subcat_map[pv] = tb_map.get(pv, 'Top&BottomWear')
        except Exception as e:
            print(f"Warning: Could not load top&bottomwear PV list: {e}")
    return pv_list, pv_subcat_map


AP_DEFAULT_CONFIG = {
    "brands": {},
    "biz_cat_id": "BCAT-139439",
    "biz_cat_name": "Apparels and Fashion",
    "relationship": "Parent",
    "catalog_status": "ACTIVE",
    "status_remark": "Ready to Launch",
    "tax_master_status": "active",
    "gst_cgst": 50,
    "gst_sgst": 50,
    "gst_igst": 0,
    "country_of_origin": "India",
    "product_condition": "Fresh",
    "manufacturing_year": "2026",
    "discovery_cat": "DISCAT-135530",
}

AP_DUMP_COL_HINTS = {
    'sku':            ['SKU_ID', 'SKU ID', 'ChildSKU', 'SKU'],
    'article':        ['ARTICLE_NUMBER', 'Article Number', 'Article Code', 'MODEL_NAME'],
    'image':          ['imageURL1', 'ImageURL1', 'imageURL1 *'],
    'image2':         ['imageURL2', 'ImageURL2'],
    'image3':         ['imageURL3', 'ImageURL3'],
    'image4':         ['imageURL4', 'ImageURL4'],
    'image5':         ['imageURL5', 'ImageURL5'],
    'image6':         ['imageURL6', 'ImageURL6'],
    'vertical':       ['Product Vertical', 'ProductVertical', 'SubType', 'Subtype'],
    'brand':          ['Brand Name', 'Brand', 'brandName'],
    'mrp':            ['MRP', 'MRP *'],
    'sp':             ['SellingPrice', 'Selling Price', 'SellingPrice *'],
    'moq':            ['MOQ', 'MOQ *'],
    'color':          ['PRODUCT_COLOR', 'Product Color', 'Color'],
    'product_desc':   ['Product Description', 'productDescription'],
    'hsn':            ['hsnCode', 'HSN', 'HSN Code'],
    'gst':            ['gstPercentage', 'GST', 'GST *'],
    'weight':         ['PRODUCT_WEIGHT_IN_KG', 'Product Weight', 'Weight'],
    'dims':           ['Product Dimension (LxBxH)', 'Product Dimension'],
    'packing':        ['PACKAGING_TYPE', 'Packaging Type'],
    'country':        ['COUNTRY_OF_ORIGIN', 'Country of Origin'],
    'set_name':       ['SET_NAME', 'Set Name', 'Set Of'],
    'quantity':       ['Quantity', 'Set Count', 'SET_COUNT'],
    'sizes':          ['AVAILABLE_SIZES', 'Available Sizes', 'Sizes'],
    'closure':        ['CLOSURE_TYPE', 'Closure Type'],
    'fabric':         ['FABRIC_MATERIAL', 'Fabric Material', 'Fabric'],
    'fit':            ['FIT', 'Fit'],
    'gender':         ['GENDER', 'Gender'],
    'length':         ['Top Length', 'Length', 'Top_Length'],
    'neck_type':      ['NECK_TYPE', 'Neck Type', 'Neck'],
    'pattern':        ['PATTERN', 'Pattern'],
    'sleeve_type':    ['SLEEVE_TYPE', 'Sleeve Type', 'Sleeve'],
    'product_type':   ['PRODUCT_TYPE', 'Product Type'],
    'core_brand':     ['CORE_BRAND', 'Core Brand'],
    'age_group':      ['AGE_GROUP', 'Age Group'],
    'bottom_closure': ['Bottom Closure', 'BOTTOM_CLOSURE *'],
    'bottom_color': ['Bottom Color', 'BOTTOM_COLOR *'],
    'bottom_fabric': ['Bottom_Fabric_Material', 'Bottom Fabric Material', 'BOTTOM_FABRIC_MATERIAL *'],
    'bottom_length': ['Bottom Length', 'BOTTOM_LENGTH *'],
    'bottom_type': ['Bottom Size', 'Bottom Type', 'BOTTOM_TYPE *', 'BOTTOM_SIZE *'],
    'bottom_pattern': ['Bottom Pattern', 'BOTTOM_PATTERN *'],
    'bottom_size': ['Bottom Size', 'BOTTOM_SIZE *'],
    'rise': ['Rise', 'RISE *'],
    'distress': ['Distress', 'DISTRESS *'],
    'stretchability': ['Stretchability', 'STRETCHABILITY'],
    'fade': ['Fade', 'FADE *'],
    'jeans_style': ['Jeans Style', 'JEANS_STYLE *'],
    'style': ['Style', 'STYLE *'],
    'dupatta_included': ['Dupatta Included', 'DUPATTA_INCLUDED'],
    'top_closure': ['Top Closure', 'TOP_CLOSURE *'],
    'top_color': ['Top Color', 'TOP_COLOR *'],
    'top_fabric': ['Top_Fabric_Material', 'Top Fabric Material', 'TOP_FABRIC_MATERIAL *'],
    'top_length': ['Top Length', 'TOP_LENGTH *'],
    'top_size': ['Top Size', 'TOP_SIZE *'],
    'top_type': ['Top Type', 'TOP_TYPE *'],
    'hemline': ['Hemline', 'HEMLINE'],
    'shape': ['Shape', 'SHAPE'],
    'occasion': ['Occasion', 'OCCASION *'],
    'number_of_pockets': ['Number of Pockets', 'NUMBER_OF_POCKETS'],
    'manufacturer': ['Manufacturer', 'MANUFACTURER'],
}

AP_BASE_COL_HINTS = {
    'article': ['ARTICLE_NUMBER', 'Article Number', 'Article Code'],
    'sku':     ['SKU_ID', 'SKU ID', 'ChildSKU'],
}

def get_ap_template_wb_for_subtype(subtype):
    path, sheet = AP_SUBTYPE_SOURCE_PATH.get(
        subtype, (AP_TEMPLATE_PATH, 'AF - PV Templates')
    )
    try:
        wb_src = load_workbook(path)
        ws_src = wb_src[sheet]
        hdr_row = AP_SUBTYPE_HEADER_ROW.get(subtype, 2)
        headers = [ws_src.cell(hdr_row, c).value for c in range(1, ws_src.max_column + 1)]
        while headers and headers[-1] is None:
            headers.pop()
    except Exception as e:
        print(f"Warning: Could not load AP template for {subtype}: {e}")
        headers = []
    headers = apply_header_renames(headers)
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = 'AF - PV Templates'
    for ci, h in enumerate(headers, 1):
        ws_new.cell(1, ci).value = h
    apply_dropdown_validations(ws_new, headers, DROPDOWN_MAP)
    return wb_new, headers

def _ap_join(parts, sep=' '):
    """Join non-empty parts with separator."""
    out = [str(p).strip() for p in parts if p is not None and str(p).strip() not in ('', 'nan', 'None', 'NaN')]
    return sep.join(out)

def build_ap_title(brand, gender, fabric, neck_type, sleeve_type, pattern, product_type, color):
    """Build AP title: Brand + Gender + Fabric + Neck + Sleeve + Pattern + Product Type, Color"""
    core = _ap_join([brand, gender, fabric, neck_type, sleeve_type, pattern, product_type])
    if color:
        return f"{core}, {color}"
    return core

def build_ap_internal_title(brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, set_name, set_details):
    """Build AP internal title: Brand, Article Number, ... , Set of N (size/qty)"""
    core = _ap_join([brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type])
    if color:
        core = f"{core}, {color}"
    if set_name:
        return f"{core}, {set_name} ({set_details})"
    return core

# ── Per-Sub-Category title formulas ──────────────────────────────
# Keyed by the "Sub Category" column header in the 'Product Vertical List'
# sheet (e.g. 'TopWears', 'InnerWears'). Add a new function + dict entry here
# whenever a new Sub Category needs its own title formula. Any Sub Category
# not listed below falls back to the generic formula.

def title_ap_topwears(brand, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, length=''):
    return build_ap_title(brand, gender, fabric, neck_type, sleeve_type, pattern, product_type, color)

def internal_title_ap_topwears(brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, set_name, set_details, length=''):
    return build_ap_internal_title(brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, set_name, set_details)

def title_ap_innerwears(brand, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, length=''):
    return build_ap_title(brand, gender, fabric, neck_type, sleeve_type, pattern, product_type, color)

def internal_title_ap_innerwears(brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, set_name, set_details, length=''):
    return build_ap_internal_title(brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, set_name, set_details)

def title_ap_bottomwears(brand, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, length=''):
    core = _ap_join([brand, gender, fabric, length, pattern, product_type])
    if color:
        return f"{core}, {color}"
    return core

def internal_title_ap_bottomwears(brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, set_name, set_details, length=''):
    core = _ap_join([brand, article, gender, fabric, length, pattern, product_type])
    if color:
        core = f"{core}, {color}"
    if set_name:
        return f"{core}, {set_name} ({set_details})"
    return core

def title_ap_topbottomwear(brand, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, length=''):
    return build_ap_title(brand, gender, fabric, neck_type, sleeve_type, pattern, product_type, color)

def internal_title_ap_topbottomwear(brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, set_name, set_details, length=''):
    return build_ap_internal_title(brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, set_name, set_details)

AP_TITLE_BUILDERS = {
    'TopWears':      title_ap_topwears,
    'InnerWears':    title_ap_innerwears,
    'BottomWear':    title_ap_bottomwears,
    'Top&BottomWear': title_ap_topbottomwear,
}
AP_INTERNAL_TITLE_BUILDERS = {
    'TopWears':      internal_title_ap_topwears,
    'InnerWears':    internal_title_ap_innerwears,
    'BottomWear':    internal_title_ap_bottomwears,
    'Top&BottomWear': internal_title_ap_topbottomwear,
}

def build_ap_titles(subcategory, brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, set_name, set_details, length=''):
    """Dispatch to the right title/internalTitle formula based on Sub Category."""
    title_fn    = AP_TITLE_BUILDERS.get(subcategory, build_ap_title)
    internal_fn = AP_INTERNAL_TITLE_BUILDERS.get(subcategory)
    title = title_fn(brand, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, length)
    if internal_fn:
        internal_title = internal_fn(brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, set_name, set_details, length)
    else:
        internal_title = build_ap_internal_title(brand, article, gender, fabric, neck_type, sleeve_type, pattern, product_type, color, set_name, set_details)
    return title, internal_title

def build_ap_set_details(sizes_str, quantity_str):
    """
    Build set details from AVAILABLE_SIZES and Quantity.
    Input: sizes='M, L, XL, 2XL', quantity='4' or '1' 
    Output: 'M/1, L/1, XL/1, 2XL/1' (distribute quantity equally)
    Or if quantity is total, divide by number of sizes
    """
    if not sizes_str:
        return '', '', ''
    
    sizes = [s.strip() for s in str(sizes_str).split(',') if s.strip()]
    if not sizes:
        return '', '', ''
    
    qty_str = str(quantity_str).strip() if quantity_str else '1'
    qty_match = re.search(r'\d+', qty_str)
    total_qty = int(qty_match.group()) if qty_match else 1
    
    # If total quantity > number of sizes, divide equally
    # If total quantity == number of sizes, 1 each
    # If set count is 1, each size gets 1
    per_size = max(1, total_qty // max(len(sizes), 1))
    
    # Actually, looking at examples: "Set of 4" with sizes M,L,XL,2XL → "S/1, M/1, L/1, XL/1"
    # So each size gets 1 when set_count=1, or quantity is per-size
    
    # Let's use the quantity as per-size quantity based on set_name parsing
    set_count = total_qty
    
    # Build set_details: S/1, M/1, L/1
    details = [f"{s}/{set_count}" for s in sizes]
    # Build set_description: 1pcs of S, 1pcs of M
    desc = [f"{set_count}pcs of {s}" for s in sizes]
    
    return ', '.join(details), ', '.join(desc), ', '.join(sizes)

def fill_ap_template(ws, headers, rows_df, col_map, subtype, existing_articles, existing_skus):
    tcol = {h: i+1 for i, h in enumerate(headers) if h}
    _ap_cfg = get_ap_config_from_disk()
    brands_dict = normalize_brands(_ap_cfg.get('brands', {}))
    fallback_brand, fallback_id = ('', '')
    if brands_dict:
        fallback_brand, fallback_id = next(iter(brands_dict.items()))
    st_data = AP_SUBTYPE_MAP.get(subtype, {})
    skipped, filled = [], []
    
    for _, drow in rows_df.iterrows():
        brand, brand_id = get_brand_info(drow, col_map, brands_dict)
        if not brand and fallback_brand:
            brand = fallback_brand
            brand_id = fallback_id
        
        sku_raw = safe(drow.get(col_map.get('sku', ''), ''))
        art_raw = safe(drow.get(col_map.get('article', ''), ''))
        article = art_raw if art_raw else sku_raw
        
        if article.upper() in existing_articles or sku_raw.upper() in existing_skus:
            skipped.append({'sku': sku_raw, 'article': article, 'reason': 'Already exists in base data'})
            continue
        
        filled_count = len(filled) + 1
        row_idx = filled_count + 1
        
        # Extract fields from dump
        mrp = drow.get(col_map.get('mrp', ''), '')
        sp = drow.get(col_map.get('sp', ''), '')
        moq = drow.get(col_map.get('moq', ''), 1)
        color = safe(drow.get(col_map.get('color', ''), ''))
        weight = drow.get(col_map.get('weight', ''), '')
        dim_raw = safe(drow.get(col_map.get('dims', ''), ''))
        hsn = drow.get(col_map.get('hsn', ''), '')
        gst = drow.get(col_map.get('gst', ''), 5)
        packing = safe(drow.get(col_map.get('packing', ''), '')) or 'Loose Packing'
        country = safe(drow.get(col_map.get('country', ''), '')) or _ap_cfg['country_of_origin']
        prod_desc = safe(drow.get(col_map.get('product_desc', ''), ''))
        set_name_raw = safe(drow.get(col_map.get('set_name', ''), ''))
        quantity_raw = safe(drow.get(col_map.get('quantity', ''), ''))
        sizes_raw = safe(drow.get(col_map.get('sizes', ''), ''))
        closure = safe(drow.get(col_map.get('closure', ''), ''))
        fabric = safe(drow.get(col_map.get('fabric', ''), ''))
        fit = safe(drow.get(col_map.get('fit', ''), ''))
        gender = safe(drow.get(col_map.get('gender', ''), ''))
        length = safe(drow.get(col_map.get('length', ''), ''))
        neck_type = safe(drow.get(col_map.get('neck_type', ''), ''))
        pattern = safe(drow.get(col_map.get('pattern', ''), ''))
        sleeve_type = safe(drow.get(col_map.get('sleeve_type', ''), ''))
        product_type = safe(drow.get(col_map.get('product_type', ''), ''))
        core_brand = safe(drow.get(col_map.get('core_brand', ''), ''))
        age_group = safe(drow.get(col_map.get('age_group', ''), ''))
        
        # BottomWear + mixed-set fields
        bottom_closure      = safe(drow.get(col_map.get('bottom_closure', ''), ''))
        bottom_color        = safe(drow.get(col_map.get('bottom_color', ''), ''))
        bottom_fabric       = safe(drow.get(col_map.get('bottom_fabric', ''), ''))
        bottom_length       = safe(drow.get(col_map.get('bottom_length', ''), ''))
        bottom_type         = safe(drow.get(col_map.get('bottom_type', ''), ''))
        bottom_pattern      = safe(drow.get(col_map.get('bottom_pattern', ''), ''))
        bottom_size         = safe(drow.get(col_map.get('bottom_size', ''), ''))
        rise                = safe(drow.get(col_map.get('rise', ''), ''))
        distress            = safe(drow.get(col_map.get('distress', ''), ''))
        stretchability      = safe(drow.get(col_map.get('stretchability', ''), ''))
        fade                = safe(drow.get(col_map.get('fade', ''), ''))
        jeans_style         = safe(drow.get(col_map.get('jeans_style', ''), ''))
        style               = safe(drow.get(col_map.get('style', ''), ''))
        dupatta_included    = safe(drow.get(col_map.get('dupatta_included', ''), ''))
        top_closure         = safe(drow.get(col_map.get('top_closure', ''), ''))
        top_color           = safe(drow.get(col_map.get('top_color', ''), ''))
        top_fabric          = safe(drow.get(col_map.get('top_fabric', ''), ''))
        top_length          = safe(drow.get(col_map.get('top_length', ''), ''))
        top_size            = safe(drow.get(col_map.get('top_size', ''), ''))
        top_type            = safe(drow.get(col_map.get('top_type', ''), ''))
        hemline             = safe(drow.get(col_map.get('hemline', ''), ''))
        shape               = safe(drow.get(col_map.get('shape', ''), ''))
        occasion            = safe(drow.get(col_map.get('occasion', ''), ''))
        number_of_pockets   = safe(drow.get(col_map.get('number_of_pockets', ''), ''))
        manufacturer        = safe(drow.get(col_map.get('manufacturer', ''), ''))
        
        img_url = safe(drow.get(col_map.get('image', ''), ''))
        img2_url = safe(drow.get(col_map.get('image2', ''), ''))
        img3_url = safe(drow.get(col_map.get('image3', ''), ''))
        img4_url = safe(drow.get(col_map.get('image4', ''), ''))
        img5_url = safe(drow.get(col_map.get('image5', ''), ''))
        img6_url = safe(drow.get(col_map.get('image6', ''), ''))
        
        # Parse set_name for count: "Set of 1" → 1, "Set of 4" → 4
        set_count = 1
        if set_name_raw:
            m = re.search(r'\d+', set_name_raw)
            if m:
                set_count = int(m.group())
        
        # Build set details
        set_details, set_desc, avail_sizes = build_ap_set_details(sizes_raw, quantity_raw or str(set_count))
        
                # Build title and internal title, routed by the PV's Sub Category
        # (TopWears / InnerWears / future categories) from Product Vertical List sheet
        title_subcategory = AP_PV_SUBCATEGORY.get(subtype, 'TopWears')
        # Fall back to Top/Bottom-specific columns if the generic top-level
        # fields are empty (BottomWear -> bottom_*, Top&BottomWear -> top_* then bottom_*)
        fabric_for_title  = fabric or top_fabric or bottom_fabric
        length_for_title  = length or top_length or bottom_length
        title, internal_title = build_ap_titles(
            title_subcategory, brand, article, gender, fabric_for_title, neck_type, sleeve_type, pattern, product_type,
            color, set_name_raw or f'Set of {set_count}', set_details, length_for_title
        )
        
        # Parse dimensions
        L, B, H = parse_lbh(dim_raw)
        
        # Parse numeric values
        try: mrp = float(mrp) if str(mrp).strip() not in ('', 'nan') else ''
        except: mrp = ''
        try: sp = float(sp) if str(sp).strip() not in ('', 'nan') else ''
        except: sp = ''
        try: hsn = int(float(hsn)) if str(hsn).strip() not in ('', 'nan') else ''
        except: hsn = ''
        gst = parse_gst_percentage(gst, default=5)
        try: moq = int(float(moq)) if str(moq).strip() not in ('', 'nan') else 1
        except: moq = 1
        try: weight = float(weight) if str(weight).strip() not in ('', 'nan') else ''
        except: weight = ''
        
        # Product code = article
        product_code = article
        
        row_data = {
            'Category *': st_data.get('Category *', 'Apparels & Fashion'),
            'SubCategory *': st_data.get('SubCategory *', ''),
            'CategoryType *': st_data.get('CategoryType *', ''),
            'SubType': subtype,
            'PVID *': st_data.get('PVID *', ''),
            'BusinessCategoryId *': _ap_cfg['biz_cat_id'],
            'BusinessCategoryName *': _ap_cfg['biz_cat_name'],
            'ProductCode *': product_code,
            'Relationship *': _ap_cfg['relationship'],
            'ParentProductId *': sku_raw,
            'ChildSKU *': sku_raw,
            'MRP *': mrp,
            'SellingPrice *': sp,
            'MOQ *': moq,
            'title *': title,
            'internalTitle *': internal_title,
            'brandId *': brand_id,
            'brandName *': brand,
            'imageURL1 *': img_url,
            'imageURL2': img2_url,
            'imageURL3': img3_url,
            'imageURL4': img4_url,
            'imageURL5': img5_url,
            'imageURL6': img6_url,
            'catalogStatus *': _ap_cfg['catalog_status'],
            'statusRemark': _ap_cfg['status_remark'],
            'discoveryCategoryIds': st_data.get('discoveryCategoryIds', _ap_cfg['discovery_cat']),
            'productDescription *': prod_desc if prod_desc else '',
            'PRODUCT_IDENTIFIER *': 'Set',
            'SET_NAME *': set_name_raw or f'Set of {set_count}',
            'SET_COUNT *': set_count,
            'PACK_NAME *': 'Pack of 1',
            'PACK_OF *': 1,
            'IS_COMBO *': 'yes',
            'AVAILABLE_SIZES *': avail_sizes,
            'SET_DETAILS *': set_details,
            'SET_DESCRIPTION *': set_desc,
            'PRODUCT_COLOR *': color,
            'ARTICLE_NUMBER *': article,
            'MODEL_NAME *': article,
            'PRODUCT_CONDITION *': _ap_cfg['product_condition'],
            'UNIT_OF_MEASUREMENT_SINGULAR *': 'Peice',
            'UNIT_OF_MEASUREMENT_PLURAL *': 'Peices',
            'UNIT_OF_MEASUREMENT_SINGULAR_ABBREVIATION *': 'Pc',
            'UNIT_OF_MEASUREMENT_PLURAL_ABBREVIATION *': 'Pcs',
            'SELLER_SKU_ID *': sku_raw,
            'PACKAGING_TYPE *': packing,
            'DESCRIPTION': '',
            'CLOSURE_TYPE *': closure,
            'COUNTRY_OF_ORIGIN *': country,
            'EAN': '',
            'IMPORTED_BY': '',
            'KEY_FEATURES': '',
            'MANUFACTURING_YEAR': _ap_cfg['manufacturing_year'],
            'PRODUCT_LENGTH *': L,
            'PRODUCT_BREADTH *': B,
            'PRODUCT_HEIGHT *': H,
            'PRODUCT_DIMENSION_UOM *': 'cm',
            'PRODUCT_TYPE *': product_type,
            'PRODUCT_WEIGHT_IN_KG *': weight,
            'PRODUCT_MANUFACTURING_CITY': '',
            'PRODUCT_MANUFACTURING_STATE': '',
            'FABRIC_MATERIAL *': fabric,
            'FIT *': fit,
            'GENDER *': gender,
            'GSM': '',
            'HEMLINE': '',
            'LENGTH *': length,
            'MANUFACTURER': '',
            'NECK_TYPE *': neck_type,
            'NUMBER_OF_POCKETS': '',
            'OCCASION': '',
            'PATTERN *': pattern,
            'SLEEVE_TYPE *': sleeve_type,
            'CORE_BRAND *': core_brand,
            'PV_SPECIFIC_SIZES *': '#',
            'hsnCode *': hsn,
            'gstPercentage *': gst,
            'cgstShare *': _ap_cfg['gst_cgst'],
            'sgstShare *': _ap_cfg['gst_sgst'],
            'igstShare *': _ap_cfg['gst_igst'],
            'cess': '',
            'sinTax': '',
            'vatPercentage': '',
            'otherCess': '',
            'validityPeriodStartDate': '',
            'validityPeriodEndDate': '',
            'declarationForm': '',
            'taxMasterStatus': _ap_cfg['tax_master_status'],
            'AGE_GROUP *': age_group,
            'BOTTOM_CLOSURE *': bottom_closure,
            'BOTTOM_COLOR *': bottom_color,
            'BOTTOM_FABRIC_MATERIAL *': bottom_fabric,
            'BOTTOM_LENGTH *': bottom_length,
            'BOTTOM_TYPE *': bottom_type,
            'BOTTOM_PATTERN *': bottom_pattern,
            'BOTTOM_SIZE *': bottom_size,
            'RISE *': rise,
            'RISE': rise,
            'DISTRESS *': distress,
            'STRETCHABILITY': stretchability,
            'FADE *': fade,
            'JEANS_STYLE *': jeans_style,
            'STYLE *': style,
            'DUPATTA_INCLUDED': dupatta_included,
            'TOP_CLOSURE *': top_closure,
            'TOP_COLOR *': top_color,
            'TOP_FABRIC_MATERIAL *': top_fabric,
            'TOP_LENGTH *': top_length,
            'TOP_SIZE *': top_size,
            'TOP_TYPE *': top_type,
            'HEMLINE': hemline,
            'SHAPE': shape,
            'OCCASION': occasion,
            'OCCASION *': occasion,
            'NUMBER_OF_POCKETS': number_of_pockets,
            'MANUFACTURER': manufacturer,
        }
        
        for col_name, val in row_data.items():
            if val is None or str(val) in ('None', ''):
                continue
            norm_name = _norm_header(col_name)
            target_col = tcol.get(col_name) or tcol.get(norm_name)
            if target_col is None:
                renamed = HEADER_RENAME_MAP.get(col_name) or HEADER_RENAME_MAP.get(norm_name)
                if renamed:
                    target_col = tcol.get(renamed) or tcol.get(_norm_header(renamed))
            if target_col:
                ws.cell(row=row_idx, column=target_col).value = val
        
        filled.append({'sku': sku_raw, 'article': article})
    
    return len(filled), skipped


# ═══════════════════════════════════════════════════════════════
# IN-MEMORY FILE STORAGE
# ═══════════════════════════════════════════════════════════════
FILE_STORE = {}

# ═══════════════════════════════════════════════════════════════
# DAILY ACTIVITY REPORT (emailed via the same FillForge SMTP mailbox)
# Triggered externally by Vercel Cron hitting /api/cron/daily_report
# ═══════════════════════════════════════════════════════════════

REPORT_TO_EMAIL = os.environ.get('REPORT_TO_EMAIL', 'alibin.hussain@jumbotail.com')
CRON_SECRET     = os.environ.get('CRON_SECRET', '')

CATEGORY_ACTION_MAP = {
    'fw_catalog_generated': 'Footwear',
    'ce_catalog_generated': 'Consumer Electronics',
    'ap_catalog_generated': 'Apparel & Fashion',
}

def send_email_smtp(to_email, subject, html_content, from_email=None):
    """Send an email using the same SMTP setup already used for OTPs."""
    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = from_email or SMTP_USER
        msg['To']      = to_email
        msg.attach(MIMEText(html_content, 'html'))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.starttls()
            s.login(SMTP_USER, SMTP_PASSWORD)
            s.sendmail(SMTP_USER, to_email, msg.as_string())
        return True, 'sent'
    except Exception as e:
        return False, str(e)


def _parse_log_details(details):
    """Extract subtypes list and filled count from a details string like:
    "subtypes=['Sneakers', 'Sandals'] filled=12 skipped=2" """
    subtypes, filled = [], 0
    m_sub = re.search(r"subtypes=\[(.*?)\]", details or '')
    if m_sub:
        subtypes = [s.strip().strip("'\"") for s in m_sub.group(1).split(',') if s.strip()]
    m_fill = re.search(r"filled=(\d+)", details or '')
    if m_fill:
        filled = int(m_fill.group(1))
    return subtypes, filled


def build_daily_report_data():
    """Pull last-24h activity from activity_logs and structure it for the report."""
    if not DATABASE_URL:
        return None, 'DATABASE_URL not configured'
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute('''
            SELECT ts, email, action, details FROM activity_logs
            WHERE ts >= NOW() - INTERVAL '24 hours'
            ORDER BY ts DESC
        ''')
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        return None, str(e)

    logins = []
    generations = []

    for r in rows:
        email  = r['email']
        action = r['action']
        ts     = str(r['ts'])
        if action == 'login_success':
            logins.append({'email': email, 'ts': ts})
        elif action in CATEGORY_ACTION_MAP:
            subtypes, filled = _parse_log_details(r['details'])
            generations.append({
                'email': email, 'ts': ts,
                'category': CATEGORY_ACTION_MAP[action],
                'pvs': subtypes, 'filled': filled,
            })

    unique_users = sorted(set(l['email'] for l in logins))
    return {
        'logins': logins,
        'unique_user_count': len(unique_users),
        'unique_users': unique_users,
        'generations': generations,
    }, None


def build_daily_report_html(data):
    login_rows = ''.join(
        f"<tr><td style='padding:6px 10px;border:1px solid #eee;'>{l['ts']}</td>"
        f"<td style='padding:6px 10px;border:1px solid #eee;'>{l['email']}</td></tr>"
        for l in data['logins']
    ) or "<tr><td colspan='2' style='padding:6px 10px;border:1px solid #eee;color:#888;'>No logins in the last 24 hours</td></tr>"

    gen_rows = ''.join(
        f"<tr><td style='padding:6px 10px;border:1px solid #eee;'>{g['ts']}</td>"
        f"<td style='padding:6px 10px;border:1px solid #eee;'>{g['email']}</td>"
        f"<td style='padding:6px 10px;border:1px solid #eee;'>{g['category']}</td>"
        f"<td style='padding:6px 10px;border:1px solid #eee;'>{', '.join(g['pvs']) or '-'}</td>"
        f"<td style='padding:6px 10px;border:1px solid #eee;text-align:center;'>{g['filled']}</td></tr>"
        for g in data['generations']
    ) or "<tr><td colspan='5' style='padding:6px 10px;border:1px solid #eee;color:#888;'>No catalogs generated in the last 24 hours</td></tr>"

    today_str = datetime.now(ZoneInfo('Asia/Kolkata')).strftime('%d %b %Y')

    return f"""
    <div style="font-family:Arial,sans-serif;max-width:720px;margin:auto;padding:24px;">
      <h2 style="color:#E87722;">FillForge — Daily Activity Report ({today_str})</h2>
      <p><b>Unique users logged in (24h):</b> {data['unique_user_count']}</p>

      <h3>Logins</h3>
      <table style="border-collapse:collapse;width:100%;font-size:13px;">
        <tr style="background:#f5f5f5;">
          <th style="padding:6px 10px;border:1px solid #eee;text-align:left;">Time</th>
          <th style="padding:6px 10px;border:1px solid #eee;text-align:left;">User Email</th>
        </tr>
        {login_rows}
      </table>

      <h3 style="margin-top:24px;">Catalog Generations</h3>
      <table style="border-collapse:collapse;width:100%;font-size:13px;">
        <tr style="background:#f5f5f5;">
          <th style="padding:6px 10px;border:1px solid #eee;text-align:left;">Time</th>
          <th style="padding:6px 10px;border:1px solid #eee;text-align:left;">User Email</th>
          <th style="padding:6px 10px;border:1px solid #eee;text-align:left;">Category</th>
          <th style="padding:6px 10px;border:1px solid #eee;text-align:left;">PV(s) Generated</th>
          <th style="padding:6px 10px;border:1px solid #eee;text-align:center;">Rows Filled</th>
        </tr>
        {gen_rows}
      </table>
      <p style="color:#999;font-size:12px;margin-top:20px;">Automated report from FillForge.</p>
    </div>
    """


def send_daily_report():
    data, err = build_daily_report_data()
    if err:
        print(f'daily_report error: {err}')
        send_email_smtp(REPORT_TO_EMAIL, 'FillForge Daily Report — ERROR',
                         f'<p>Could not generate report: {err}</p>')
        return False, err
    html = build_daily_report_html(data)
    ok, msg = send_email_smtp(REPORT_TO_EMAIL, 'FillForge — Daily Activity Report', html)
    write_log('system', 'daily_report_sent' if ok else 'daily_report_failed', msg)
    print(f'daily report send: {ok} — {msg}')
    return ok, msg


@app.route('/api/cron/daily_report', methods=['GET', 'POST'])
def cron_daily_report():
    """Called by Vercel Cron once a day. Protected by a shared secret,
    since cron calls have no browser session/cookie."""
    auth_header = request.headers.get('Authorization', '')
    if not CRON_SECRET or auth_header != f'Bearer {CRON_SECRET}':
        return jsonify({'error': 'Unauthorized'}), 401
    ok, msg = send_daily_report()
    if ok:
        return jsonify({'status': 'ok', 'message': 'Daily report sent'})
    return jsonify({'status': 'error', 'message': msg}), 500


# ═══════════════════════════════════════════════════════════════
# ROUTES
# ═══════════════════════════════════════════════════════════════
def _load_json(path):
    if not os.path.exists(path): return {}
    try:
        with open(path, 'r') as f: return json.load(f)
    except: return {}

def _save_json(path, data):
    with open(path, 'w') as f: json.dump(data, f)

def send_otp(email):
    otp    = str(secrets.randbelow(900000) + 100000)
    expiry = datetime.utcnow() + timedelta(minutes=OTP_EXPIRY_MINUTES)
    if DATABASE_URL:
        try:
            conn = get_db()
            cur  = conn.cursor()
            cur.execute(
                'INSERT INTO otps (email, otp, expiry) VALUES (%s, %s, %s) '
                'ON CONFLICT (email) DO UPDATE SET otp=%s, expiry=%s',
                (email, otp, expiry, otp, expiry)
            )
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            print(f'send_otp db error: {e}')

    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'FillForge Login OTP'
        msg['From']    = SMTP_USER
        msg['To']      = email
        html = f"""
        <div style="font-family:Arial,sans-serif;max-width:480px;margin:auto;padding:32px;
                    border:1px solid #eee;border-radius:12px;">
          <h2 style="color:#E87722;">FillForge Login</h2>
          <p>Your one-time password is:</p>
          <div style="font-size:36px;font-weight:bold;letter-spacing:8px;
                      color:#1a1a1a;padding:16px 0;">{otp}</div>
          <p style="color:#888;font-size:13px;">Valid for {OTP_EXPIRY_MINUTES} minutes.
             Do not share this with anyone.</p>
        </div>"""
        msg.attach(MIMEText(html, 'html'))
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as s:
            s.starttls()
            s.login(SMTP_USER, SMTP_PASSWORD)
            s.sendmail(SMTP_USER, email, msg.as_string())
        return True, 'OTP sent'
    except Exception as e:
        return False, str(e)

def verify_otp(email, otp_input):
    if not DATABASE_URL: return False, 'Database not configured'
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute('SELECT otp, expiry FROM otps WHERE email=%s', (email,))
        row  = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return False, 'No OTP found for this email'
        if datetime.utcnow() > row['expiry']:
            cur.execute('DELETE FROM otps WHERE email=%s', (email,))
            conn.commit()
            cur.close(); conn.close()
            return False, 'OTP expired'
        if row['otp'] != str(otp_input).strip():
            cur.close(); conn.close()
            return False, 'Invalid OTP'
        cur.execute('DELETE FROM otps WHERE email=%s', (email,))
        conn.commit()
        cur.close(); conn.close()
        return True, 'OK'
    except Exception as e:
        return False, str(e)

def create_session(email):
    token  = secrets.token_hex(32)
    expiry = datetime.utcnow() + timedelta(days=SESSION_EXPIRY_DAYS)
    if DATABASE_URL:
        try:
            conn = get_db()
            cur  = conn.cursor()
            # Drop any older sessions for this email first — avoids repeated
            # test logins piling up as separate "active" entries later.
            cur.execute('DELETE FROM sessions WHERE email=%s', (email,))
            cur.execute(
                'INSERT INTO sessions (token, email, expiry, last_seen) VALUES (%s, %s, %s, NOW())',
                (token, email, expiry)
            )
            conn.commit()
            cur.close()
            conn.close()
        except Exception as e:
            print(f'create_session error: {e}')
    return token

def validate_session(token):
    if not token or not DATABASE_URL: return None
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute('SELECT email, expiry FROM sessions WHERE token=%s', (token,))
        row  = cur.fetchone()
        if not row:
            cur.close(); conn.close()
            return None
        if datetime.utcnow() > row['expiry']:
            cur.execute('DELETE FROM sessions WHERE token=%s', (token,))
            conn.commit()
            cur.close(); conn.close()
            return None
        cur.execute('UPDATE sessions SET last_seen = NOW() WHERE token=%s', (token,))
        conn.commit()
        cur.close()
        conn.close()
        return row['email']
    except Exception as e:
        print(f'validate_session error: {e}')
        return None

def require_auth(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.cookies.get('ff_session')
        if not validate_session(token):
            if request.path == '/' or not request.path.startswith('/'):
                return redirect('/login')
            return jsonify({'error': 'Unauthorized', 'redirect': '/login'}), 401
        return f(*args, **kwargs)
    return decorated
# ── Global error handlers ─────────────────────────────────────
@app.errorhandler(400)
def bad_request(e):
    return jsonify({'error': 'Bad request', 'details': str(e)}), 400

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Not found', 'details': str(e)}), 404

@app.errorhandler(500)
def server_error(e):
    import traceback
    return jsonify({'error': 'Internal server error', 'details': str(e), 'trace': traceback.format_exc()}), 500

@app.errorhandler(Exception)
def unhandled_exception(e):
    import traceback
    return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/login')
def login_page():
    return render_template('login.html')

@app.route('/auth/send_otp', methods=['POST'])
def auth_send_otp():
    email = (request.json or {}).get('email', '').strip().lower()
    if not email or '@' not in email:
        return jsonify({'error': 'Invalid email'}), 400
    ok, msg = send_otp(email)
    if ok:
        write_log(email, 'otp_sent')
        return jsonify({'status': 'ok'})
    return jsonify({'error': f'Failed to send OTP: {msg}'}), 500

@app.route('/auth/verify_otp', methods=['POST'])
def auth_verify_otp():
    data  = request.json or {}
    email = data.get('email', '').strip().lower()
    otp   = data.get('otp', '').strip()
    ok, msg = verify_otp(email, otp)
    if ok:
        token = create_session(email)
        write_log(email, 'login_success')
        resp = jsonify({'status': 'ok'})
        resp.set_cookie('ff_session', token,
                        max_age=SESSION_EXPIRY_DAYS * 86400,
                        httponly=True, samesite='Lax')
        return resp
    write_log(email, 'login_failed', msg)
    return jsonify({'error': msg}), 401

@app.route('/auth/logout', methods=['POST'])
def auth_logout():
    token = request.cookies.get('ff_session')
    if token and DATABASE_URL:
        try:
            conn = get_db()
            cur  = conn.cursor()
            cur.execute('DELETE FROM sessions WHERE token=%s', (token,))
            conn.commit()
            cur.close()
            conn.close()
        except: pass
    resp = jsonify({'status': 'ok'})
    resp.set_cookie('ff_session', '', expires=0)
    return resp

@app.route('/auth/me')
def auth_me():
    token = request.cookies.get('ff_session')
    email = validate_session(token)
    if email:
        return jsonify({'email': email})
    return jsonify({'error': 'Not logged in'}), 401

@app.route('/auth/active_users')
def active_users():
    """Presence pills: who has an active FillForge session in the last 2 minutes."""
    if not DATABASE_URL or not PSYCOPG2_AVAILABLE:
        return jsonify({'users': []})
    try:
        conn = get_db()
        cur  = conn.cursor()
        cur.execute('''
            SELECT DISTINCT email FROM sessions
            WHERE last_seen >= NOW() - INTERVAL '2 minutes'
            ORDER BY email
        ''')
        emails = [r['email'] for r in cur.fetchall()]
        cur.close(); conn.close()
        return jsonify({'users': emails})
    except Exception as e:
        print(f'active_users error: {e}')
        return jsonify({'users': [], 'error': str(e)})

@app.route('/')
@require_auth
def index():
    try:
        return render_template('index.html')
    except Exception as e:
        try:
            html_path = os.path.join(os.path.dirname(__file__), 'templates', 'index.html')
            if not os.path.exists(html_path):
                html_path = os.path.join(os.path.dirname(__file__), 'index.html')
            with open(html_path, 'r', encoding='utf-8') as f:
                html = f.read()
            html = html.replace("{{ user_email|default('', true) }}", '')
            return html
        except Exception as e2:
            import traceback
            return f"<h1>Template Error</h1><pre>{traceback.format_exc()}</pre>", 500

@app.route('/tools/ticket-closer')
@require_auth
def ticket_closer():
    return render_template('ticket_closer.html')

@app.route('/subtypes')
def get_subtypes():
    return jsonify({'subtypes': PV_LIST})

@app.route('/ce_subtypes')
def get_ce_subtypes():
    return jsonify({'subtypes': CE_PV_LIST})

# ── Stub routes for removed modules (AP, TS, CE5) ─────────────
# These keep the old frontend from crashing with JSON parse errors.

@app.route('/ap_categories')
def get_ap_categories():
    # AP_SUBTYPE_MAP (built from the 'AF - PV Templates' sheet) is what auto-detect
    # and catalog generation actually rely on, and is proven reliable. AP_PV_LIST
    # comes from a separate 'Product Vertical List' sheet that can fail to parse
    # independently — merge both so the dropdown never comes back empty as long
    # as the main template loaded correctly.
    combined = list(AP_SUBTYPE_MAP.keys())
    for pv in AP_PV_LIST:
        if pv not in combined:
            combined.append(pv)
    return jsonify({'subtypes': combined})
    
@app.route('/ts_categories')
def get_ts_categories():
    return jsonify({'categories': []})

@app.route('/ce5_categories')
def get_ce5_categories():
    return jsonify({'categories': []})

@app.route('/ap_config', methods=['GET'])
def ap_config_get_route():
    return jsonify(get_ap_config_from_disk())

@app.route('/ap_config', methods=['POST'])
def update_ap_config():
    cfg = get_ap_config_from_disk()
    data = request.json
    if 'brands' in data:
        data['brands'] = normalize_brands(data['brands'])
    cfg.update(data)
    _save_config(AP_CONFIG_PATH, cfg)
    write_log(validate_session(request.cookies.get('ff_session')) or 'anonymous', 'ap_config_updated', f"brands={cfg.get('brands')}")
    return jsonify({'status': 'ok'})

@app.route('/ts_config', methods=['GET'])
def ts_config_get_route():
    return jsonify({'brands': {}, 'biz_cat_id': '', 'biz_cat_name': 'Toys & Sports',
                    'catalog_status': 'ACTIVE', 'status_remark': 'Ready to Launch',
                    'tax_master_status': 'active', 'gst_cgst': 50, 'gst_sgst': 50,
                    'gst_igst': 0, 'country_of_origin': 'India', 'product_condition': 'Fresh',
                    'manufacturing_year': '2026', 'discovery_cat': '', 'relationship': 'Parent',
                    'pv_config': {}})

@app.route('/ts_config', methods=['POST'])
def update_ts_config():
    return jsonify({'status': 'ok'})

@app.route('/ce5_config', methods=['GET'])
def ce5_config_get_route():
    return jsonify({'brands': {}})

@app.route('/ce5_config', methods=['POST'])
def update_ce5_config():
    return jsonify({'status': 'ok'})

@app.route('/detect_ap_categories', methods=['POST'])
def detect_ap_categories():
    try:
        dump_file = request.files.get('dump')
        if not dump_file:
            return jsonify({'categories': []})
        
        # Read all sheets from the uploaded Excel
        xl = pd.ExcelFile(io.BytesIO(dump_file.read()))
        frames = []
        for sname in xl.sheet_names:
            try: frames.append(xl.parse(sname))
            except: pass
        all_dump = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        
        # Normalize column names (remove extra spaces)
        all_dump = normalize_df_columns(all_dump)
        
        # Map input columns to expected fields using AP_DUMP_COL_HINTS
        col_map = build_col_map(all_dump, AP_DUMP_COL_HINTS)
        vert_col = col_map.get('vertical')  # This finds "Product Vertical" column
        
        if vert_col and vert_col in all_dump.columns:
            # Get all unique values from Product Vertical column
            found = [str(v).strip() for v in all_dump[vert_col].dropna().unique()
                     if str(v).strip() not in ('nan', 'None', '')]
            # Filter to only those that exist in our template
            matched = [v for v in found if v in AP_SUBTYPE_MAP]
            return jsonify({'verticals': matched, 'all_found': found})
        
        return jsonify({'verticals': [], 'all_found': []})
    except Exception as e:
        return jsonify({'verticals': [], 'error': str(e)})

@app.route('/detect_ts_categories', methods=['POST'])
def detect_ts_categories():
    return jsonify({'categories': [], 'all_found': []})

@app.route('/detect_ce5_categories', methods=['POST'])
def detect_ce5_categories():
    return jsonify({'categories': [], 'all_found': []})

@app.route('/process_ap', methods=['POST'])
def process_ap():
    """Apparel & Fashion catalog processor. Generates a filled AF - PV Template .xlsx."""
    try:
        # ── 1. Parse subtypes from form ─────────────────────────────
        subtypes_raw = request.form.get('subtypes', '')
        try:
            subtypes = json.loads(subtypes_raw)
        except:
            subtypes = [s.strip() for s in subtypes_raw.split(',') if s.strip()]

        # ── 2. Parse inline config (optional) ───────────────────────
        inline_cfg_raw = request.form.get('ap_config', '')
        if inline_cfg_raw:
            try:
                inline_cfg = json.loads(inline_cfg_raw)
                if inline_cfg.get('brands'):
                    inline_cfg['brands'] = normalize_brands(inline_cfg['brands'])
                disk_cfg = get_ap_config_from_disk()
                disk_cfg.update(inline_cfg)
                _save_config(AP_CONFIG_PATH, disk_cfg)
            except Exception as e:
                print(f'inline ap_config parse error: {e}')

        # ── 3. Get uploaded files ───────────────────────────────────
        base_file = request.files.get('base_data')
        dump_file = request.files.get('dump')

        # ── 4. Validate inputs ──────────────────────────────────────
        if not subtypes:
            return jsonify({'error': 'Please select at least one Product Vertical'}), 400
        if not dump_file:
            return jsonify({'error': 'Dump / listing file is required'}), 400
        for st in subtypes:
            if st not in AP_SUBTYPE_MAP:
                return jsonify({'error': f'Product Vertical "{st}" not found in AP template'}), 400

        # ── 5. Read dump file ───────────────────────────────────────
        dump_bytes = dump_file.read()
        xl = pd.ExcelFile(io.BytesIO(dump_bytes))
        frames = []
        for sname in xl.sheet_names:
            try:
                frames.append(xl.parse(sname))
            except:
                pass
        all_dump = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        if all_dump.empty:
            return jsonify({'error': 'Could not read any data from dump file'}), 400

        all_dump = normalize_df_columns(all_dump)

        # ── 6. Map columns ──────────────────────────────────────────
        col_map = build_col_map(all_dump, AP_DUMP_COL_HINTS)
        vert_col = col_map.get('vertical')

        # ── 7. Read base data for duplicate check ───────────────────
        existing_articles, existing_skus = set(), set()
        if base_file:
            bxl = pd.ExcelFile(io.BytesIO(base_file.read()))
            for sname in bxl.sheet_names:
                try:
                    bdf = bxl.parse(sname)
                    bcol = build_col_map(bdf, AP_BASE_COL_HINTS)
                    if 'article' in bcol:
                        existing_articles |= set(bdf[bcol['article']].dropna().astype(str).str.strip().str.upper())
                    if 'sku' in bcol:
                        existing_skus |= set(bdf[bcol['sku']].dropna().astype(str).str.strip().str.upper())
                except:
                    pass

        # ── 8. Process each subtype ─────────────────────────────────
        results, all_skipped, grand_filled = [], [], 0
        preview_rows, preview_cols = [], []

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for subtype in subtypes:
                # Filter rows for this subtype
                if vert_col and vert_col in all_dump.columns:
                    mask = all_dump[vert_col].astype(str).str.strip().str.lower() == subtype.lower()
                    filtered = all_dump[mask].copy()
                    if filtered.empty:
                        mask2 = all_dump[vert_col].astype(str).str.lower().str.contains(
                            re.escape(subtype.lower()), na=False)
                        filtered = all_dump[mask2].copy()
                    if filtered.empty:
                        filtered = all_dump.copy()
                else:
                    filtered = all_dump.copy()

                # Get blank template for this subtype
                wb, headers = get_ap_template_wb_for_subtype(subtype)
                ws = wb.active

                # Fill template with data
                filled, skipped = fill_ap_template(
                    ws, headers, filtered, col_map, subtype, existing_articles, existing_skus
                )
                all_skipped.extend(skipped)
                grand_filled += filled

                # Save to ZIP
                safe_st = re.sub(r"[^\w\s-]", "", subtype).replace(" ", "_")
                fname = f'ap_filled_{safe_st}.xlsx'
                xls_buf = io.BytesIO()
                wb.save(xls_buf)
                zout.writestr(fname, xls_buf.getvalue())
                results.append({
                    'subtype': subtype,
                    'filled': filled,
                    'skipped': len(skipped),
                    'filename': fname
                })

                # Build preview columns (first time only)
                if not preview_cols:
                    pcols = [
                        'title *', 'ChildSKU *', 'ARTICLE_NUMBER *', 'MRP *', 'SellingPrice *',
                        'PRODUCT_COLOR *', 'AVAILABLE_SIZES *', 'SET_DETAILS *',
                        'SET_COUNT *', 'FABRIC_MATERIAL *', 'NECK_TYPE *', 'PATTERN *',
                        'SLEEVE_TYPE *', 'hsnCode *'
                    ]
                    preview_cols = [c for c in pcols if c in headers]

                # Build preview rows (max 50)
                for r in range(2, min(filled + 2, 52)):
                    rdata = {}
                    for c in preview_cols:
                        if c in headers:
                            rdata[c] = ws.cell(r, headers.index(c) + 1).value
                    if any(v for v in rdata.values()):
                        preview_rows.append({**rdata, '_subtype': subtype})

        # ── 9. Prepare output ───────────────────────────────────────
        zip_buf.seek(0)

        if len(subtypes) == 1:
            safe_st = re.sub(r"[^\w\s-]", "", subtypes[0]).replace(" ", "_")
            out_name = f'ap_filled_{safe_st}.xlsx'
            out_ext = '.xlsx'
            with zipfile.ZipFile(io.BytesIO(zip_buf.getvalue())) as zin:
                out_bytes = zin.read(results[0]['filename'])
        else:
            out_name = 'ap_filled_templates.zip'
            out_ext = '.zip'
            out_bytes = zip_buf.getvalue()

        # Store in memory for download
        file_token = ''.join(random.choices(string.ascii_letters + string.digits, k=32))
        FILE_STORE[file_token] = {
            'bytes': out_bytes,
            'filename': out_name,
            'ext': out_ext,
            'created': time.time()
        }

        # ── 10. Log and return ──────────────────────────────────────
        write_log(
            validate_session(request.cookies.get('ff_session')) or 'anonymous',
            'ap_catalog_generated',
            f'subtypes={subtypes} filled={grand_filled} skipped={len(all_skipped)}'
        )

        return jsonify({
            'status': 'ok',
            'grand_filled': grand_filled,
            'grand_skipped': len(all_skipped),
            'results': results,
            'skipped_details': all_skipped[:50],
            'preview': preview_rows,
            'preview_cols': preview_cols,
            'download_token': file_token,
            'filename': out_name,
            'is_zip': len(subtypes) > 1,
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/process_ts', methods=['POST'])
def process_ts():
    return jsonify({'error': 'Toys & Sports module is not available in this version.'}), 400

@app.route('/process_ce5', methods=['POST'])
def process_ce5():
    return jsonify({'error': 'CE 5-File module is not available in this version.'}), 400

@app.route('/config', methods=['GET'])
def config_get_route():
    return jsonify(get_config())

@app.route('/ce_config', methods=['GET'])
def ce_config_get_route():
    return jsonify(get_ce_config_from_disk())

@app.route('/config', methods=['POST'])
def update_config():
    cfg  = get_config()
    data = request.json
    if 'brands' in data:
        data['brands'] = normalize_brands(data['brands'])
    cfg.update(data)
    _save_config(CONFIG_PATH, cfg)
    write_log(validate_session(request.cookies.get('ff_session')) or 'anonymous', 'config_updated', f"brands={cfg.get('brands')}")
    return jsonify({'status': 'ok'})

@app.route('/ce_config', methods=['POST'])
def update_ce_config():
    cfg  = get_ce_config_from_disk()
    data = request.json
    if 'brands' in data:
        data['brands'] = normalize_brands(data['brands'])
    cfg.update(data)
    _save_config(CE_CONFIG_PATH, cfg)
    write_log(validate_session(request.cookies.get('ff_session')) or 'anonymous', 'ce_config_updated', f"brands={cfg.get('brands')}")
    return jsonify({'status': 'ok'})

@app.route('/logs')
def get_logs():
    return jsonify({'logs': read_logs(500)})

@app.route('/logs/export')
def export_logs():
    logs = read_logs(2000)
    if not logs:
        return jsonify({'error': 'No logs found'}), 404
    df = pd.DataFrame(logs)
    # Rename columns for clarity
    df.columns = [c.upper() for c in df.columns]
    if 'TS' in df.columns:
        df.rename(columns={'TS': 'TIMESTAMP', 'EMAIL': 'USER EMAIL',
                            'ACTION': 'ACTION', 'DETAILS': 'DETAILS'}, inplace=True)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Activity Logs')
        ws = writer.sheets['Activity Logs']
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    buf.seek(0)
    from datetime import datetime
    fname = f'FillForge_ActivityLogs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/detect_verticals', methods=['POST'])
def detect_verticals():
    try:
        dump_file = request.files.get('dump')
        if not dump_file:
            return jsonify({'verticals': []})
        xl     = pd.ExcelFile(io.BytesIO(dump_file.read()))
        frames = []
        for sname in xl.sheet_names:
            try: frames.append(xl.parse(sname))
            except: pass
        all_dump = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        all_dump = normalize_df_columns(all_dump)
        col_map  = build_col_map(all_dump, DUMP_COL_HINTS)
        vert_col = col_map.get('vertical')
        if vert_col and vert_col in all_dump.columns:
            found   = [str(v).strip() for v in all_dump[vert_col].dropna().unique()
                       if str(v).strip() not in ('nan','None','')]
            matched = [v for v in found if v in SUBTYPE_MAP]
            return jsonify({'verticals': matched, 'all_found': found})
        return jsonify({'verticals': [], 'all_found': []})
    except Exception as e:
        return jsonify({'verticals': [], 'error': str(e)})

@app.route('/detect_ce_verticals', methods=['POST'])
def detect_ce_verticals():
    try:
        dump_file = request.files.get('dump')
        if not dump_file:
            return jsonify({'verticals': []})
        xl     = pd.ExcelFile(io.BytesIO(dump_file.read()))
        frames = []
        for sname in xl.sheet_names:
            try: frames.append(xl.parse(sname))
            except: pass
        all_dump = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        col_map  = build_col_map(all_dump, CE_DUMP_COL_HINTS)
        vert_col = col_map.get('vertical')
        if vert_col and vert_col in all_dump.columns:
            found   = [str(v).strip() for v in all_dump[vert_col].dropna().unique()
                       if str(v).strip() not in ('nan','None','')]
            matched = [v for v in found if v in CE_SUBTYPE_MAP]
            return jsonify({'verticals': matched, 'all_found': found})
        return jsonify({'verticals': [], 'all_found': []})
    except Exception as e:
        return jsonify({'verticals': [], 'error': str(e)})

@app.route('/process', methods=['POST'])
def process():
    """Footwear catalog processor. Generates a filled PV Template .xlsx."""
    try:
        subtypes_raw = request.form.get('subtypes', '')
        try:    subtypes = json.loads(subtypes_raw)
        except: subtypes = [s.strip() for s in subtypes_raw.split(',') if s.strip()]

        inline_cfg_raw = request.form.get('config', '')
        if inline_cfg_raw:
            try:
                inline_cfg = json.loads(inline_cfg_raw)
                if inline_cfg.get('brands'):
                    inline_cfg['brands'] = normalize_brands(inline_cfg['brands'])
                disk_cfg = get_config()
                disk_cfg.update(inline_cfg)
                _save_config(CONFIG_PATH, disk_cfg)
            except Exception as e:
                print(f'inline config parse error: {e}')

        base_file = request.files.get('base_data')
        dump_file = request.files.get('dump')

        if not subtypes:
            return jsonify({'error': 'Please select at least one SubType'}), 400
        if not dump_file:
            return jsonify({'error': 'Dump / listing file is required'}), 400
        for st in subtypes:
            if st not in SUBTYPE_MAP:
                return jsonify({'error': f'SubType "{st}" not found in template'}), 400

        dump_bytes = dump_file.read()
        xl         = pd.ExcelFile(io.BytesIO(dump_bytes))
        frames     = []
        for sname in xl.sheet_names:
            try: frames.append(xl.parse(sname))
            except: pass
        all_dump = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        if all_dump.empty:
            return jsonify({'error': 'Could not read any data from dump file'}), 400

        col_map  = build_col_map(all_dump, DUMP_COL_HINTS)
        vert_col = col_map.get('vertical')

        existing_articles, existing_skus = set(), set()
        if base_file:
            bxl = pd.ExcelFile(io.BytesIO(base_file.read()))
            for sname in bxl.sheet_names:
                try:
                    bdf  = bxl.parse(sname)
                    bcol = build_col_map(bdf, CE_BASE_COL_HINTS)
                    if 'article' in bcol:
                        existing_articles |= set(bdf[bcol['article']].dropna().astype(str).str.strip().str.upper())
                    if 'sku' in bcol:
                        existing_skus |= set(bdf[bcol['sku']].dropna().astype(str).str.strip().str.upper())
                except: pass

        results, all_skipped, grand_filled = [], [], 0
        preview_rows, preview_cols = [], []

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for subtype in subtypes:
                if vert_col and vert_col in all_dump.columns:
                    mask     = all_dump[vert_col].astype(str).str.strip().str.lower() == subtype.lower()
                    filtered = all_dump[mask].copy()
                    if filtered.empty:
                        mask2    = all_dump[vert_col].astype(str).str.lower().str.contains(re.escape(subtype.lower()), na=False)
                        filtered = all_dump[mask2].copy()
                    if filtered.empty:
                        filtered = all_dump.copy()
                else:
                    filtered = all_dump.copy()

                wb, headers = get_template_wb_for_subtype(subtype)
                ws = wb.active
                filled, skipped = fill_template(
                    ws, headers, filtered, col_map, subtype, existing_articles, existing_skus
                )
                all_skipped.extend(skipped)
                grand_filled += filled

                safe_st = re.sub(r"[^\w\s-]", "", subtype).replace(" ", "_")
                fname   = f'filled_{safe_st}.xlsx'
                xls_buf = io.BytesIO()
                wb.save(xls_buf)
                zout.writestr(fname, xls_buf.getvalue())
                results.append({'subtype': subtype, 'filled': filled,
                                 'skipped': len(skipped), 'filename': fname})

                if not preview_cols:
                    pcols = ['title *','ChildSKU *','ARTICLE_NUMBER *','MRP *','SellingPrice *',
                             'PRODUCT_COLOR *','AVAILABLE_SIZES *','SET_DETAILS *',
                             'SET_COUNT *','FOOTWEAR_TYPE *','hsnCode *']
                    preview_cols = [c for c in pcols if c in headers]
                for r in range(2, min(filled + 2, 52)):
                    rdata = {}
                    for c in preview_cols:
                        if c in headers:
                            rdata[c] = ws.cell(r, headers.index(c)+1).value
                    if any(v for v in rdata.values()):
                        preview_rows.append({**rdata, '_subtype': subtype})

        zip_buf.seek(0)
        if len(subtypes) == 1:
            safe_st  = re.sub(r"[^\w\s-]", "", subtypes[0]).replace(" ", "_")
            out_name = f'filled_{safe_st}.xlsx'
            out_ext  = '.xlsx'
            with zipfile.ZipFile(io.BytesIO(zip_buf.getvalue())) as zin:
                out_bytes = zin.read(results[0]['filename'])
        else:
            out_name  = 'filled_footwear_templates.zip'
            out_ext   = '.zip'
            out_bytes = zip_buf.getvalue()

        file_token = ''.join(random.choices(string.ascii_letters + string.digits, k=32))
        FILE_STORE[file_token] = {'bytes': out_bytes, 'filename': out_name,
                                   'ext': out_ext, 'created': time.time()}

        write_log(validate_session(request.cookies.get('ff_session')) or 'anonymous', 'fw_catalog_generated',
                  f'subtypes={subtypes} filled={grand_filled} skipped={len(all_skipped)}')

        return jsonify({
            'status': 'ok', 'grand_filled': grand_filled,
            'grand_skipped': len(all_skipped), 'results': results,
            'skipped_details': all_skipped[:50], 'preview': preview_rows,
            'preview_cols': preview_cols, 'download_token': file_token,
            'filename': out_name, 'is_zip': len(subtypes) > 1,
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/process_ce', methods=['POST'])
def process_ce():
    try:
        subtypes_raw = request.form.get('subtypes', '')
        try:    subtypes = json.loads(subtypes_raw)
        except: subtypes = [s.strip() for s in subtypes_raw.split(',') if s.strip()]

        inline_cfg_raw = request.form.get('ce_config', '')
        if inline_cfg_raw:
            try:
                inline_cfg = json.loads(inline_cfg_raw)
                if inline_cfg.get('brands'):
                    inline_cfg['brands'] = normalize_brands(inline_cfg['brands'])
                try:
                    disk_cfg = get_ce_config_from_disk()
                    disk_cfg.update(inline_cfg)
                    _save_config(CE_CONFIG_PATH, disk_cfg)
                except: pass
            except Exception as e:
                print(f'inline ce_config parse error: {e}')

        base_file = request.files.get('base_data')
        dump_file = request.files.get('dump')

        if not subtypes:
            return jsonify({'error': 'Please select at least one SubType'}), 400
        if not dump_file:
            return jsonify({'error': 'Dump / listing file is required'}), 400
        for st in subtypes:
            if st not in CE_SUBTYPE_MAP:
                return jsonify({'error': f'SubType "{st}" not found in CE template'}), 400

        dump_bytes = dump_file.read()
        xl         = pd.ExcelFile(io.BytesIO(dump_bytes))
        frames     = []
        for sname in xl.sheet_names:
            try: frames.append(xl.parse(sname))
            except: pass
        all_dump = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        if all_dump.empty:
            return jsonify({'error': 'Could not read any data from dump file'}), 400
        all_dump = normalize_df_columns(all_dump)

        col_map  = build_col_map(all_dump, CE_DUMP_COL_HINTS)
        vert_col = col_map.get('vertical')

        existing_articles, existing_skus = set(), set()
        if base_file:
            bxl = pd.ExcelFile(io.BytesIO(base_file.read()))
            for sname in bxl.sheet_names:
                try:
                    bdf  = bxl.parse(sname)
                    bcol = build_col_map(bdf, CE_BASE_COL_HINTS)
                    if 'article' in bcol:
                        existing_articles |= set(bdf[bcol['article']].dropna().astype(str).str.strip().str.upper())
                    if 'sku' in bcol:
                        existing_skus |= set(bdf[bcol['sku']].dropna().astype(str).str.strip().str.upper())
                except: pass

        results, all_skipped, grand_filled = [], [], 0
        preview_rows, preview_cols = [], []

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for subtype in subtypes:
                if vert_col and vert_col in all_dump.columns:
                    mask     = all_dump[vert_col].astype(str).str.strip().str.lower() == subtype.lower()
                    filtered = all_dump[mask].copy()
                    if filtered.empty:
                        mask2    = all_dump[vert_col].astype(str).str.lower().str.contains(re.escape(subtype.lower()), na=False)
                        filtered = all_dump[mask2].copy()
                    if filtered.empty:
                        filtered = all_dump.copy()
                else:
                    filtered = all_dump.copy()

                wb, headers = get_ce_template_wb_for_subtype(subtype)
                ws = wb.active
                filled, skipped = fill_ce_template(
                    ws, headers, filtered, col_map, subtype, existing_articles, existing_skus
                )
                all_skipped.extend(skipped)
                grand_filled += filled

                safe_st = re.sub(r"[^\w\s-]", "", subtype).replace(" ", "_")
                fname   = f'ce_filled_{safe_st}.xlsx'
                xls_buf = io.BytesIO()
                wb.save(xls_buf)
                zout.writestr(fname, xls_buf.getvalue())
                results.append({'subtype': subtype, 'filled': filled,
                                 'skipped': len(skipped), 'filename': fname})

                if not preview_cols:
                    pcols = ['title *','ChildSKU *','ARTICLE_NUMBER *','MRP *','SellingPrice *',
                             'PRODUCT_COLOR *','RAM *','INTERNAL_STORAGE *',
                             'DISPLAY_SIZE *','DISPLAY_TYPE *','hsnCode *','SET_COUNT *']
                    preview_cols = [c for c in pcols if c in headers]
                    for r in range(2, min(filled + 2, 52)):
                        rdata = {c: ws.cell(r, headers.index(c)+1).value for c in preview_cols}
                        if any(v for v in rdata.values()):
                            preview_rows.append({**rdata, '_subtype': subtype})

        zip_buf.seek(0)
        if len(subtypes) == 1:
            safe_st  = re.sub(r"[^\w\s-]", "", subtypes[0]).replace(" ", "_")
            out_name = f'ce_filled_{safe_st}.xlsx'
            out_ext  = '.xlsx'
            with zipfile.ZipFile(io.BytesIO(zip_buf.getvalue())) as zin:
                out_bytes = zin.read(results[0]['filename'])
        else:
            out_name  = 'ce_filled_templates.zip'
            out_ext   = '.zip'
            out_bytes = zip_buf.getvalue()

        file_token = ''.join(random.choices(string.ascii_letters + string.digits, k=32))
        FILE_STORE[file_token] = {'bytes': out_bytes, 'filename': out_name,
                                   'ext': out_ext, 'created': time.time()}

        write_log(validate_session(request.cookies.get('ff_session')) or 'anonymous', 'ce_catalog_generated',
                  f'subtypes={subtypes} filled={grand_filled} skipped={len(all_skipped)}')

        return jsonify({
            'status': 'ok', 'grand_filled': grand_filled,
            'grand_skipped': len(all_skipped), 'results': results,
            'skipped_details': all_skipped[:50], 'preview': preview_rows,
            'preview_cols': preview_cols, 'download_token': file_token,
            'filename': out_name, 'is_zip': len(subtypes) > 1,
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


# ── Subtype-specific template generator ─────────────────────────
def _generate_blank_template(subtype):
    """Generate a blank Excel template containing only the header row for a specific subtype."""
    if subtype not in SUBTYPE_HEADER_ROW:
        return None, f'SubType "{subtype}" not found in template'

    wb_src  = load_workbook(TEMPLATE_PATH)
    ws_src  = wb_src['PV Template']
    hdr_row = SUBTYPE_HEADER_ROW.get(subtype, 1)

    headers = []
    for c in range(1, ws_src.max_column + 1):
        val = ws_src.cell(hdr_row, c).value
        if val is not None:
            headers.append(str(val).strip())

    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = 'PV Template'

    for ci, h in enumerate(headers, 1):
        ws_new.cell(1, ci).value = h

    for ci, h in enumerate(headers, 1):
        ws_new.column_dimensions[ws_new.cell(1, ci).column_letter].width = max(len(h) + 2, 15)

    buf = io.BytesIO()
    wb_new.save(buf)
    buf.seek(0)
    return buf, None

@app.route('/download_ce_unified_template', methods=['GET'])
def download_ce_unified_template():
    category = request.args.get('category', '').strip()
    VALID_CATEGORIES = [
        'Adapters & Cables',
        'Audio Devices',
        'Memory & Storage',
        'Mobile Batteries',
        'Cases & Protectors',
        'Smartphones & Watches',
    ]
    if category not in VALID_CATEGORIES:
        return jsonify({'error': f'Invalid category: {category}'}), 400

    unified_path = os.path.join(os.path.dirname(__file__), 'Consumer_Electronics_Templates_Merged.xlsx')
    if not os.path.exists(unified_path):
        return jsonify({'error': 'Unified template file not found'}), 404

    try:
        wb_src = load_workbook(unified_path)
        if category not in wb_src.sheetnames:
            return jsonify({'error': f'Sheet "{category}" not found in unified template'}), 404

        wb_new = Workbook()
        wb_new.remove(wb_new.active)

        def copy_sheet(ws_src, wb_dest, sheet_title):
            ws_new = wb_dest.create_sheet(title=sheet_title)
            for row in ws_src.iter_rows():
                for cell in row:
                    new_cell = ws_new.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font          = copy.copy(cell.font)
                        new_cell.border        = copy.copy(cell.border)
                        new_cell.fill          = copy.copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection    = copy.copy(cell.protection)
                        new_cell.alignment     = copy.copy(cell.alignment)
            for col_letter, col_dim in ws_src.column_dimensions.items():
                ws_new.column_dimensions[col_letter].width = col_dim.width
            for row_idx, row_dim in ws_src.row_dimensions.items():
                ws_new.row_dimensions[row_idx].height = row_dim.height
            for dv in ws_src.data_validations.dataValidation:
                ws_new.add_data_validation(dv)
            return ws_new

        if 'Dropdown Reference' in wb_src.sheetnames:
            copy_sheet(wb_src['Dropdown Reference'], wb_new, 'Dropdown Reference')
            wb_new['Dropdown Reference'].sheet_state = 'hidden'

        copy_sheet(wb_src[category], wb_new, category)

        buf = io.BytesIO()
        wb_new.save(buf)
        buf.seek(0)
        safe_name = category.replace(' ', '_').replace('&', 'and')
        fname = f'CE_Template_{safe_name}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500

@app.route('/download_template/<path:vertical>')
def download_template(vertical):
    """Download blank template for a specific vertical/category."""
    try:
        if vertical == 'Footwear':
            # Return the full footwear master template
            if not os.path.exists(TEMPLATE_PATH):
                return jsonify({'error': 'Footwear template not found'}), 404
            
            wb = load_workbook(TEMPLATE_PATH)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf, as_attachment=True,
                           download_name='Footwear_Master_Template.xlsx',
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        elif vertical == 'Apparel & Fashion':
            # Return the full AP master template
            if not os.path.exists(AP_TEMPLATE_PATH):
                return jsonify({'error': 'Apparel & Fashion template not found'}), 404
            
            wb = load_workbook(AP_TEMPLATE_PATH)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf, as_attachment=True,
                           download_name='Apparel_Fashion_Master_Template.xlsx',
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        elif vertical == 'Consumer Electronics':
            # Return the full CE master template
            if not os.path.exists(CE_TEMPLATE_PATH):
                return jsonify({'error': 'Consumer Electronics template not found'}), 404
            
            wb = load_workbook(CE_TEMPLATE_PATH)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return send_file(buf, as_attachment=True,
                           download_name='Consumer_Electronics_Master_Template.xlsx',
                           mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        else:
            return jsonify({'error': f'Unknown vertical: {vertical}'}), 400
            
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/reload_templates', methods=['POST'])
@require_auth
def reload_templates():
    try:
        SUBTYPE_HEADER_ROW_new, SUBTYPE_MAP_new = _build_header_row_map()
        PV_LIST_new = load_pv_list()
        CE_SUBTYPE_HEADER_ROW_new, CE_SUBTYPE_MAP_new = _build_ce_header_row_map()
        CE_PV_LIST_new = load_ce_pv_list()
        AP_SUBTYPE_HEADER_ROW_new, AP_SUBTYPE_MAP_new = _build_ap_header_row_map()
        AP_PV_LIST_new, AP_PV_SUBCATEGORY_new = load_ap_pv_list()
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    global SUBTYPE_HEADER_ROW, SUBTYPE_MAP, PV_LIST
    global CE_SUBTYPE_HEADER_ROW, CE_SUBTYPE_MAP, CE_PV_LIST
    global AP_SUBTYPE_HEADER_ROW, AP_SUBTYPE_MAP, AP_PV_LIST, AP_PV_SUBCATEGORY
    SUBTYPE_HEADER_ROW, SUBTYPE_MAP = SUBTYPE_HEADER_ROW_new, SUBTYPE_MAP_new
    PV_LIST = PV_LIST_new
    CE_SUBTYPE_HEADER_ROW, CE_SUBTYPE_MAP = CE_SUBTYPE_HEADER_ROW_new, CE_SUBTYPE_MAP_new
    CE_PV_LIST = CE_PV_LIST_new
    AP_SUBTYPE_HEADER_ROW, AP_SUBTYPE_MAP = AP_SUBTYPE_HEADER_ROW_new, AP_SUBTYPE_MAP_new
    AP_PV_LIST, AP_PV_SUBCATEGORY = AP_PV_LIST_new, AP_PV_SUBCATEGORY_new

    write_log(validate_session(request.cookies.get('ff_session')) or 'anonymous',
              'templates_reloaded', '')
    return jsonify({'status': 'ok', 'message': 'Templates reloaded from disk'})

@app.route('/download/<token>')
def download(token):
    if '..' in token or '/' in token or '\\' in token: return 'Invalid', 400
    if token in FILE_STORE:
        file_data = FILE_STORE[token]
        ext   = file_data['ext']
        fname = request.args.get('filename', file_data['filename'])
        mtype = 'application/zip' if ext == '.zip' else \
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        write_log(validate_session(request.cookies.get('ff_session')) or 'anonymous', 'file_downloaded', fname)
        return send_file(io.BytesIO(file_data['bytes']), as_attachment=True,
                         download_name=fname, mimetype=mtype)
    tmpdir = tempfile.gettempdir()
    for ext in ['', '.zip', '.xlsx']:
        path = os.path.join(tmpdir, token + ext)
        if os.path.exists(path):
            fname = request.args.get('filename', 'filled_template' + ext)
            mtype = 'application/zip' if ext == '.zip' else \
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            write_log('anonymous', 'file_downloaded', fname)
            return send_file(path, as_attachment=True, download_name=fname, mimetype=mtype)
    return 'File not found', 404

@app.route('/test_brand')
def test_brand():
    cfg         = get_config()
    brands_raw  = cfg.get('brands', {})
    brands_dict = normalize_brands(brands_raw)
    fallback_brand, fallback_id = ('', '')
    if brands_dict:
        fallback_brand, fallback_id = next(iter(brands_dict.items()))
    return jsonify({
        'step1_raw_from_config_json': brands_raw,
        'step2_after_normalize':      brands_dict,
        'step3_fallback_brand':       fallback_brand,
        'step4_fallback_id':          fallback_id,
        'step5_config_json_exists':   os.path.exists(CONFIG_PATH),
        'step6_config_json_content':  open(CONFIG_PATH).read() if os.path.exists(CONFIG_PATH) else 'FILE NOT FOUND',
        'conclusion': 'Brand will appear in title' if fallback_brand else 'NO BRAND — brands dict is empty!',
    })

@app.route('/debug_config')
def debug_config():
    cfg    = get_config()
    ce_cfg = get_ce_config_from_disk()
    return jsonify({
        'config_path':           CONFIG_PATH,
        'ce_config_path':        CE_CONFIG_PATH,
        'config_file_exists':    os.path.exists(CONFIG_PATH),
        'ce_config_file_exists': os.path.exists(CE_CONFIG_PATH),
        'footwear_brands':       cfg.get('brands', {}),
        'ce_brands':             ce_cfg.get('brands', {}),
        'footwear_config':       cfg,
        'ce_config':             ce_cfg,
    })

@app.route('/debug_ap')
def debug_ap():
    """Diagnose why Apparel & Fashion Product Verticals might be loading empty."""
    info = {
        'ap_template_path':      AP_TEMPLATE_PATH,
        'ap_template_exists':    os.path.exists(AP_TEMPLATE_PATH),
        'ap_pv_list_count':      len(AP_PV_LIST),
        'ap_pv_list_sample':     AP_PV_LIST[:10],
        'ap_subtype_map_count':  len(AP_SUBTYPE_MAP),
    }
    if os.path.exists(AP_TEMPLATE_PATH):
        try:
            wb = load_workbook(AP_TEMPLATE_PATH)
            info['sheet_names'] = wb.sheetnames
            info['has_pv_list_sheet']       = 'Product Vertical List' in wb.sheetnames
            info['has_af_templates_sheet']  = 'AF - PV Templates' in wb.sheetnames
            if 'Product Vertical List' in wb.sheetnames:
                ws = wb['Product Vertical List']
                info['pv_sheet_dimensions'] = ws.dimensions
                info['pv_sheet_row1_headers'] = [
                    ws.cell(1, c).value for c in range(1, min(ws.max_column, 10) + 1)
                ]
        except Exception as e:
            info['workbook_read_error'] = str(e)
    return jsonify(info)


if __name__ == '__main__':
    app.run(debug=False, port=5050)
