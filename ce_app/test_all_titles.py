"""E2E verification: feed one row per SubType and inspect the generated
title + internalTitle against the CE - Mapping Logic formulas."""
import io, json, requests, openpyxl, zipfile
from openpyxl import Workbook

# (subtype, dump_row) — column names match CE_DUMP_COL_HINTS aliases.
ROWS = [
    # Feature Phones: Brand+Model+Screen Size "Display"+Product Type, Color (Condition)
    {'sku':'FP-001','Product Type':'Feature Phones','Brand':'Nokia','Model Number':'105',
     'Product Color':'Black','Screen Size':'1.8','MRP':1500,'Selling Price':1199},
    # Smart Phone: Brand+Model+Back Camera "Camera"+Type, RAM+Storage, Color (Condition)
    {'sku':'SP-001','Product Type':'Smart Phone','Brand':'Samsung','Model Number':'Galaxy A15',
     'Product Color':'Blue','Back Camera':'50MP','RAM':'6GB','Storage Capacity':'128GB',
     'MRP':20000,'Selling Price':14999},
    # Mobile Adapters & Cables: Brand+Condition+Output Voltage+Adapter Connector Type, Color
    {'sku':'AC-001','Product Type':'Mobile Adapters & Cables','Brand':'Mi','Model Number':'33W',
     'Output Voltage':'33W','Adapter Connector Type':'USB Type-C','Product Color':'White',
     'MRP':1499,'Selling Price':999},
    # Hair Trimmer: Brand+Model+Condition+Battery Type+Type, Color
    {'sku':'HT-001','Product Type':'Hair Trimmer','Brand':'Philips','Model Number':'BT3105',
     'Battery Type':'Lithium-Ion','Product Color':'Black','MRP':2499,'Selling Price':1799},
    # Speakers: Brand+Model+Condition+Speaker Type
    {'sku':'SPK-001','Product Type':'Speakers','Brand':'JBL','Model Number':'Flip 6',
     'Speaker Type':'Bluetooth','Product Color':'Red','MRP':12999,'Selling Price':9999},
    # Mobile Case & Covers: Brand+Condition+Compatible Brand+Model Name+Material+Case Type, Color
    {'sku':'CC-001','Product Type':'Mobile Case & Covers','Brand':'Spigen','Model Number':'Ultra Hybrid',
     'Compatible Brand':'Apple','Compatible Model':'iPhone 15','Material':'Polycarbonate',
     'Case Cover Type':'Back Cover','Product Color':'Crystal Clear','MRP':2499,'Selling Price':1499},
    # Earphones: Brand+Model+Condition+"Wired Earphones"
    {'sku':'EAR-001','Product Type':'Earphones','Brand':'boAt','Model Number':'Bassheads 100',
     'Product Color':'Black','MRP':999,'Selling Price':399},
    # Headsets: Brand+Model+Condition+Type
    {'sku':'HS-001','Product Type':'Headsets','Brand':'boAt','Model Number':'Rockerz 450',
     'Product Color':'Blue','MRP':2999,'Selling Price':1499},
    # Memory Cards: Brand+Storage Capacity+Memory Card Type+Type, Speed Class
    {'sku':'MC-001','Product Type':'Memory Cards','Brand':'SanDisk','Model Number':'Ultra',
     'Storage Capacity':'128GB','Memory Card Type':'microSDXC','Speed Class':'Class 10',
     'MRP':1499,'Selling Price':1099},
    # Mobile Cables: Brand+Condition+No.of Connectors+Connector Type+Type, Color
    {'sku':'CAB-001','Product Type':'Mobile Cables','Brand':'Anker','Model Number':'PowerLine III',
     'Number of Connectors':'2','Adapter Connector Type':'USB-C to Lightning',
     'Product Color':'White','MRP':1499,'Selling Price':999},
    # Mobile Holders: Brand+Condition+Type+Rotation, Color
    {'sku':'MH-001','Product Type':'Mobile Holders','Brand':'Tukzer','Model Number':'TZ-CM01',
     'Rotation':'360 Degree','Product Color':'Black','MRP':799,'Selling Price':499},
    # Screen Guards / Protectors: Brand+Condition+Guard Type+Coverage+"for"+Compatible Brand+Model
    {'sku':'SG-001','Product Type':'Screen Guards / Protectors','Brand':'AmazonBasics','Model Number':'AB-SG-15',
     'Screen Guard Type':'Tempered Glass','Coverage':'Full Screen',
     'Compatible Brand':'Apple','Compatible Model':'iPhone 15 Pro',
     'Product Color':'Clear','MRP':499,'Selling Price':199},
    # Neck Bands: Brand+Model+Condition+"Neckband"
    {'sku':'NB-001','Product Type':'Neck Bands','Brand':'Realme','Model Number':'Buds Wireless 3',
     'Product Color':'Black','MRP':1999,'Selling Price':1299},
    # Microphone: Brand+Model+Condition+Mic Type+"with"+Output Connector Type
    {'sku':'MIC-001','Product Type':'Microphone','Brand':'Shure','Model Number':'MV7',
     'Mic Type':'Dynamic','Output Connector Type':'XLR/USB','Product Color':'Black',
     'MRP':29999,'Selling Price':24999},
    # Power Bank: Brand+Condition+Battery Capacity+Type+No.of Ports+Port Type, Color
    {'sku':'PB-001','Product Type':'Power Bank','Brand':'Mi','Model Number':'HyperSonic',
     'Battery Capacity':'20000mAh','Number of Output Ports':'2','Port Type':'USB Type-C',
     'Product Color':'Black','MRP':2499,'Selling Price':1799},
    # Smart Watches: Brand+Model+Condition+Screen Size "Display"+Type, Color
    {'sku':'SW-001','Product Type':'Smart Watches','Brand':'Noise','Model Number':'Pulse 2 Max',
     'Screen Size':'1.85','Product Color':'Jet Black','MRP':2999,'Selling Price':1799},
    # TWS Ear Buds: Brand+Model+Condition+"Earbuds"
    {'sku':'TWS-001','Product Type':'TWS Ear Buds','Brand':'OnePlus','Model Number':'Nord Buds 2',
     'Product Color':'Thunder Gray','MRP':2999,'Selling Price':2299},
    # Projectors (no formula → generic fallback)
    {'sku':'PRJ-001','Product Type':'Projectors','Brand':'BenQ','Model Number':'TH685',
     'Product Color':'White','MRP':84999,'Selling Price':69999},
]

# Build a single dump file with all rows
all_cols = sorted({k for r in ROWS for k in r})
wb = Workbook(); ws = wb.active; ws.title = 'Sheet1'
ws.append(['Child SKU' if c=='sku' else c for c in all_cols])
for r in ROWS:
    ws.append([r.get(c, '') for c in all_cols])
buf = io.BytesIO(); wb.save(buf); buf.seek(0)

# POST with all SubTypes selected at once
subtypes = sorted({r['Product Type'] for r in ROWS})
files = {'dump': ('all.xlsx', buf.getvalue(),
                  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
brands = {r['Brand']: f"BRAND-{r['Brand'].upper().replace(' ','')}" for r in ROWS}
data = {'subtypes': json.dumps(subtypes),
        'ce_config': json.dumps({'brands': brands})}

r = requests.post('http://localhost:5050/process_ce', files=files, data=data, timeout=60)
print(f'\nHTTP {r.status_code}  grand_filled={r.json().get("grand_filled")}')

dl = requests.get(f'http://localhost:5050/download/{r.json()["download_token"]}', timeout=60)
z = zipfile.ZipFile(io.BytesIO(dl.content))

print(f'\n{"SubType":<30}{"title":<90}internalTitle')
print('─' * 200)
for name in sorted(z.namelist()):
    sub_wb = openpyxl.load_workbook(io.BytesIO(z.read(name)))
    sub_ws = sub_wb.active
    hdrs = [c.value for c in sub_ws[1]]
    if sub_ws.max_row < 2: continue
    row2 = [c.value for c in sub_ws[2]]
    st = row2[hdrs.index('SubType')] if 'SubType' in hdrs else '?'
    title = row2[hdrs.index('title *')] if 'title *' in hdrs else ''
    itl = row2[hdrs.index('internalTitle *')] if 'internalTitle *' in hdrs else ''
    print(f'{str(st):<30}{str(title):<90}{itl}')
