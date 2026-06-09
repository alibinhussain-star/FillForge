"""Verification test: ingest a Smart Watches input file through /process_ce."""
import io, json, requests, openpyxl
from openpyxl import Workbook

# 1) Build a tiny seller-dump file containing a Smart Watches row.
wb = Workbook()
ws = wb.active
ws.title = 'Sheet1'
headers = ['Child SKU', 'Model Number', 'Product Type', 'Brand',
           'MRP', 'Selling Price', '*Minimum Order Quantity',
           'Product Color', 'Main Image URL', 'HSN Code', 'GST',
           'Product Description', 'Product Weight',
           '*Product Dimension (LXBXH)', 'Warranty Period']
ws.append(headers)
ws.append(['SW-NOISE-PULSE-BLK-001', 'Noise Pulse 2 Max', 'Smart Watches',
           'Noise', 2999, 1799, 1, 'Jet Black',
           'https://example.com/img1.jpg', 85176200, 18,
           'Smart Watch with 1.85 inch HD display, Bluetooth 5.3 calling, '
           'heart-rate sensor, 7-day battery. USB Type-C charging.',
           '0.085', '5x4x2', '1 Year'])
ws.append(['EAR-BOAT-141-RED-002', 'Boat Rockerz 141', 'Earphones',
           'boAt', 1299, 599, 1, 'Red',
           'https://example.com/img2.jpg', 85183000, 18,
           'Wired earphones with deep bass and tangle-free cable.',
           '0.020', '15x5x3', '6 Months'])
buf = io.BytesIO(); wb.save(buf); buf.seek(0)

# 2) POST to /process_ce
files = {'dump': ('test_input.xlsx', buf.getvalue(),
                  'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
data = {'subtypes': json.dumps(['Smart Watches', 'Earphones']),
        'ce_config': json.dumps({'brands': {'Noise': 'BRAND-NOISE-01',
                                            'boAt':  'BRAND-BOAT-01'}})}
r = requests.post('http://localhost:5050/process_ce', files=files, data=data, timeout=30)
print('HTTP', r.status_code)
res = r.json()
print('Status:        ', res.get('status'))
print('Grand Filled:  ', res.get('grand_filled'))
print('Grand Skipped: ', res.get('grand_skipped'))
print('Per-SubType:')
for x in res.get('results', []):
    print(f'   - {x["subtype"]:25s} filled={x["filled"]} skipped={x["skipped"]}  file={x["filename"]}')

# 3) Download the generated file and inspect the Smart Watches row
tok = res.get('download_token')
print('\nDownload token:', tok)
dl = requests.get(f'http://localhost:5050/download/{tok}', timeout=30)
print('Download HTTP:', dl.status_code, 'bytes:', len(dl.content))

# 4) Open the zip / xlsx and print the Smart Watches sheet row 1 + row 2
import zipfile
if res.get('is_zip'):
    z = zipfile.ZipFile(io.BytesIO(dl.content))
    for n in z.namelist():
        print(f'\n=== {n} ===')
        sub_wb = openpyxl.load_workbook(io.BytesIO(z.read(n)))
        sub_ws = sub_wb.active
        hdrs = [c.value for c in sub_ws[1]]
        print('Total columns:', len([h for h in hdrs if h]))
        # Show key fields for first data row
        if sub_ws.max_row >= 2:
            row2 = [c.value for c in sub_ws[2]]
            print('Key fields:')
            for key in ['Category *','SubCategory *','CategoryType *','SubType','PVID *',
                        'ChildSKU *','title *','brandName *','MRP *','PRODUCT_COLOR *']:
                if key in hdrs:
                    print(f'   {key:25s} = {row2[hdrs.index(key)]!r}')
print('\n✅ Test complete.')
